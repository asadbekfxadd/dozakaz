from flask import Flask, request, jsonify, render_template, send_from_directory, send_file, session, Response
import os, io, re
from datetime import datetime
from werkzeug.utils import secure_filename
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import psycopg2
from psycopg2.extras import RealDictCursor
from threading import Thread
import time

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dozakaz-secret-key-2026-fallback')
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024

ALLOWED = {'.xlsx', '.xls', '.csv'}
DATABASE_URL = os.environ.get('DATABASE_URL', '')

# Power BI config
PBI_TENANT_ID = os.environ.get('PBI_TENANT_ID', '7365ebbc-66ba-4712-8260-23fe81d5c482')
PBI_CLIENT_ID = os.environ.get('PBI_CLIENT_ID', '9d271a44-ec56-4859-9015-0824bac220e5')
PBI_CLIENT_SECRET = os.environ.get('PBI_CLIENT_SECRET', '')
PBI_REPORT_ID = '0e916fd9-34dc-4a5c-bb81-4addd6ca95a7'
PBI_APP_ID = '9f8ee641-5763-44a2-a684-69199c8a9d40'
_pbi_token_cache = {'token': None, 'expires': 0}

def get_pbi_token():
    import urllib.request, urllib.parse, json as _json, time
    if _pbi_token_cache['token'] and time.time() < _pbi_token_cache['expires']:
        return _pbi_token_cache['token']
    secret = PBI_CLIENT_SECRET or os.environ.get('PBI_CLIENT_SECRET', '')
    if not secret:
        return None
    url = f'https://login.microsoftonline.com/{PBI_TENANT_ID}/oauth2/v2.0/token'
    data = urllib.parse.urlencode({
        'grant_type': 'client_credentials',
        'client_id': PBI_CLIENT_ID,
        'client_secret': secret,
        'scope': 'https://analysis.windows.net/powerbi/api/.default'
    }).encode()
    try:
        req = urllib.request.Request(url, data=data, method='POST')
        with urllib.request.urlopen(req, timeout=15) as r:
            result = _json.loads(r.read())
        token = result.get('access_token')
        expires_in = result.get('expires_in', 3600)
        _pbi_token_cache['token'] = token
        _pbi_token_cache['expires'] = time.time() + expires_in - 60
        return token
    except Exception as e:
        print(f'[PBI TOKEN] Error: {e}')
        return None

USERS = {
    'admin': {'password': '123456', 'role': 'admin', 'branch': None, 'branches': None},
    'ANNA_V': {'password': 'region1', 'role': 'regional', 'branch': None, 'branches': ['HIGH TOWN PLAZA','MAGIC CITY','NOVZA','Scopus Mall']},
    'ANNA_G': {'password': 'region2', 'role': 'regional', 'branch': None, 'branches': ['ATLAS CHIMGAN','ECO PARK','HIGH TOWN PLAZA','MALIKA','Shota Rustavely','Yunusabad gallery']},
    'ALAYSKIY': {'password': 'LI-NING1', 'role': 'user', 'branch': 'ALAYSKIY', 'branches': None},
    'ATLAS CHIMGAN': {'password': 'LI-NING1', 'role': 'user', 'branch': 'ATLAS CHIMGAN', 'branches': None},
    'ECO PARK': {'password': 'LI-NING1', 'role': 'user', 'branch': 'ECO PARK', 'branches': None},
    'Family park': {'password': 'LI-NING1', 'role': 'user', 'branch': 'Family park', 'branches': None},
    'HIGH TOWN PLAZA': {'password': 'LI-NING1', 'role': 'user', 'branch': 'HIGH TOWN PLAZA', 'branches': None},
    'M. BARAKA': {'password': 'LI-NING1', 'role': 'user', 'branch': 'M. BARAKA', 'branches': None},
    'MAGIC CITY': {'password': 'LI-NING1', 'role': 'user', 'branch': 'MAGIC CITY', 'branches': None},
    'MALIKA': {'password': 'LI-NING1', 'role': 'user', 'branch': 'MALIKA', 'branches': None},
    'NOVZA': {'password': 'LI-NING1', 'role': 'user', 'branch': 'NOVZA', 'branches': None},
    'Scopus Mall': {'password': 'LI-NING1', 'role': 'user', 'branch': 'Scopus Mall', 'branches': None},
    'Shota Rustavely': {'password': 'LI-NING1', 'role': 'user', 'branch': 'Shota Rustavely', 'branches': None},
    'TASHKENT CITY MALL': {'password': 'LI-NING1', 'role': 'user', 'branch': 'TASHKENT CITY MALL', 'branches': None},
    'UZBEGIM ANDIJAN': {'password': 'LI-NING1', 'role': 'user', 'branch': 'UZBEGIM ANDIJAN', 'branches': None},
    'Yunusabad gallery': {'password': 'LI-NING1', 'role': 'user', 'branch': 'Yunusabad gallery', 'branches': None},
    'SKLAD': {'password': 'sklad2026', 'role': 'warehouse', 'branch': 'WMS', 'branches': None},
    'DMITRIY': {'password': 'Dm!LN2026$', 'role': 'admin', 'branch': None, 'branches': None},
}

BRANCHES = ['ALAYSKIY','ATLAS CHIMGAN','ECO PARK','Family park','HIGH TOWN PLAZA',
            'M. BARAKA','MAGIC CITY','MALIKA','NOVZA','Scopus Mall',
            'Shota Rustavely','TASHKENT CITY MALL','UZBEGIM ANDIJAN','Yunusabad gallery']

FLAGMANS = ['TASHKENT CITY MALL','ATLAS CHIMGAN','ALAYSKIY','Shota Rustavely']

YADISK_TOKEN = os.environ.get('YADISK_TOKEN', '35a4110c9f5a4729ba8a54cf978276f4')
YADISK_PUBLIC_KEY = 'https://disk.yandex.ru/d/uJxKSsp_PRLemQ'
YADISK_FOLDER = '/06-White Background Pics'
_photo_cache = {}
_photo_url_cache = {}

# Auto-sync config
SYNC_TOKEN = os.environ.get('SYNC_YADISK_TOKEN', 'y0__wgBEMqV35IIGNuWAyCth5vyF4mF73xoNyGCf7QJ3fjdgKw2YKVk')
SYNC_FOLDER = '/li-ning-sync/остатки'
SYNC_FILENAME = 'остатки для сайта.xlsx'
SYNC_INTERVAL_HOURS = 2
_last_sync = None

def get_db():
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    return conn

def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute('''CREATE TABLE IF NOT EXISTS orders (
        id SERIAL PRIMARY KEY,
        branch TEXT NOT NULL,
        responsible TEXT NOT NULL,
        date TEXT NOT NULL,
        priority TEXT DEFAULT 'Обычный',
        note TEXT DEFAULT '',
        filename TEXT,
        original_name TEXT,
        status TEXT DEFAULT 'Новая',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS order_items (
        id SERIAL PRIMARY KEY,
        order_id INTEGER REFERENCES orders(id),
        article TEXT,
        name TEXT,
        size TEXT,
        qty INTEGER,
        wms_stock INTEGER DEFAULT 0
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS catalog (
        id SERIAL PRIMARY KEY,
        article TEXT,
        name TEXT,
        size TEXT,
        wms_stock INTEGER DEFAULT 0,
        abc TEXT DEFAULT 'C',
        sold INTEGER DEFAULT 0,
        season TEXT DEFAULT '',
        category TEXT DEFAULT '',
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS branch_stock (
        id SERIAL PRIMARY KEY,
        article TEXT,
        size TEXT,
        branch TEXT,
        qty INTEGER DEFAULT 0
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS top_products (
        id SERIAL PRIMARY KEY,
        article TEXT,
        category TEXT,
        sold INTEGER,
        stock INTEGER,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS sales (
        id SERIAL PRIMARY KEY,
        season TEXT,
        article TEXT,
        category TEXT,
        branch TEXT,
        sale_date DATE,
        qty INTEGER DEFAULT 1,
        price NUMERIC DEFAULT 0,
        amount NUMERIC DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS schlopka_sessions (
        id SERIAL PRIMARY KEY,
        name TEXT NOT NULL,
        filename TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        created_by TEXT,
        ready_for_pickup BOOLEAN DEFAULT FALSE,
        ready_at TIMESTAMP
    )''')
    cur.execute('ALTER TABLE schlopka_sessions ADD COLUMN IF NOT EXISTS ready_for_pickup BOOLEAN DEFAULT FALSE')
    cur.execute('ALTER TABLE schlopka_sessions ADD COLUMN IF NOT EXISTS ready_at TIMESTAMP')
    cur.execute("ALTER TABLE schlopka_items ADD COLUMN IF NOT EXISTS branch_ready BOOLEAN DEFAULT FALSE")
    cur.execute("ALTER TABLE schlopka_items ADD COLUMN IF NOT EXISTS branch_taken BOOLEAN DEFAULT FALSE")
    cur.execute("ALTER TABLE catalog ADD COLUMN IF NOT EXISTS discount INTEGER DEFAULT 0")
    cur.execute('''CREATE TABLE IF NOT EXISTS schlopka_items (
        id SERIAL PRIMARY KEY,
        session_id INTEGER REFERENCES schlopka_sessions(id) ON DELETE CASCADE,
        article TEXT NOT NULL,
        name TEXT NOT NULL,
        size TEXT NOT NULL,
        branch TEXT NOT NULL,
        qty INTEGER DEFAULT 1,
        note TEXT DEFAULT '',
        status TEXT DEFAULT 'Не собран',
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    cur.execute("ALTER TABLE schlopka_items ADD COLUMN IF NOT EXISTS note TEXT DEFAULT ''")
    cur.execute('CREATE INDEX IF NOT EXISTS idx_schlopka_session ON schlopka_items(session_id)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_schlopka_branch ON schlopka_items(session_id, branch)')
    cur.execute('''CREATE TABLE IF NOT EXISTS transfers (
        id SERIAL PRIMARY KEY,
        from_branch TEXT NOT NULL,
        to_branch TEXT NOT NULL,
        article TEXT NOT NULL,
        size TEXT NOT NULL,
        qty INTEGER NOT NULL,
        status TEXT DEFAULT 'Новая',
        note TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    # Indexes for performance
    cur.execute('CREATE INDEX IF NOT EXISTS idx_catalog_article ON catalog(article)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_catalog_abc ON catalog(abc)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_branch_stock_article ON branch_stock(article, branch)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_sales_article ON sales(article)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_sales_date ON sales(sale_date)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_sales_branch ON sales(branch)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_sales_art_branch ON sales(article, branch)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_sales_art_date ON sales(article, sale_date)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_orders_branch ON orders(branch)')
    conn.commit()
    cur.close()
    conn.close()

os.makedirs('uploads', exist_ok=True)
os.makedirs('static', exist_ok=True)
init_db()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'username' not in session:
            return jsonify({'error': 'Unauthorized'}), 401
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'username' not in session:
            return jsonify({'error': 'Unauthorized'}), 401
        if session.get('role') != 'admin':
            return jsonify({'error': 'Forbidden'}), 403
        return f(*args, **kwargs)
    return decorated

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/manifest.json')
def manifest():
    return send_from_directory('.', 'manifest.json', mimetype='application/manifest+json')

@app.route('/sw.js')
def service_worker():
    return send_from_directory('.', 'sw.js', mimetype='application/javascript')

@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory('static', filename)

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    user = USERS.get(username)
    if not user or user['password'] != password:
        return jsonify({'error': 'Неверный логин или пароль'}), 401
    session['username'] = username
    session['role'] = user['role']
    session['branch'] = user.get('branch')
    session['branches'] = user.get('branches')
    return jsonify({'role': user['role'], 'branch': user.get('branch'), 'branches': user.get('branches'), 'username': username})

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'ok': True})

@app.route('/api/me')
def me():
    if 'username' not in session:
        return jsonify({'logged_in': False})
    return jsonify({'logged_in': True, 'role': session['role'], 'branch': session['branch'], 'branches': session.get('branches'), 'username': session['username']})

@app.route('/api/orders', methods=['GET'])
@login_required
def get_orders():
    branch = request.args.get('branch', '')
    status = request.args.get('status', '')
    responsible = request.args.get('responsible', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    conn = get_db(); cur = conn.cursor()
    q = 'SELECT * FROM orders WHERE 1=1'
    params = []
    if session['role'] == 'user':
        q += ' AND branch=%s'; params.append(session['branch'])
    elif session['role'] == 'regional':
        bs = session.get('branches', [])
        if branch and branch in bs:
            q += ' AND branch=%s'; params.append(branch)
        elif bs:
            q += ' AND branch = ANY(%s)'; params.append(bs)
    else:
        if branch: q += ' AND branch=%s'; params.append(branch)
    if status: q += ' AND status=%s'; params.append(status)
    if responsible: q += ' AND responsible ILIKE %s'; params.append(f'%{responsible}%')
    if date_from: q += ' AND date >= %s'; params.append(date_from)
    if date_to: q += ' AND date <= %s'; params.append(date_to)
    q += ' ORDER BY created_at DESC'
    cur.execute(q, params)
    rows = cur.fetchall()
    cur.close(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/orders/<int:oid>/items', methods=['GET'])
@login_required
def get_order_items(oid):
    conn = get_db(); cur = conn.cursor()
    cur.execute('SELECT * FROM order_items WHERE order_id=%s', (oid,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return jsonify([dict(r) for r in rows])

def upload_to_yadisk(file_path, filename):
    '''Upload file to Yandex Disk /li-ning-orders/ folder'''
    import urllib.request, urllib.parse, json as _json
    try:
        folder = '/li-ning-orders'
        # Create folder if not exists
        folder_url = ('https://cloud-api.yandex.net/v1/disk/resources?'
                     + urllib.parse.urlencode({'path': folder}))
        req = urllib.request.Request(folder_url, method='PUT',
                                    headers={'Authorization': f'OAuth {SYNC_TOKEN}'})
        try: urllib.request.urlopen(req, timeout=5)
        except: pass
        # Get upload URL
        upload_url = ('https://cloud-api.yandex.net/v1/disk/resources/upload?'
                     + urllib.parse.urlencode({'path': f'{folder}/{filename}', 'overwrite': 'true'}))
        req2 = urllib.request.Request(upload_url,
                                     headers={'Authorization': f'OAuth {SYNC_TOKEN}'})
        with urllib.request.urlopen(req2, timeout=10) as r:
            href = _json.loads(r.read()).get('href', '')
        if not href:
            return False
        # Upload file
        with open(file_path, 'rb') as f:
            req3 = urllib.request.Request(href, data=f.read(), method='PUT')
            urllib.request.urlopen(req3, timeout=30)
        return True
    except Exception as e:
        print(f'[YADISK UPLOAD] Error: {e}')
        return False

def generate_order_excel(branch, items):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Заказ'
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 10
    ws.row_dimensions[1].height = 6
    ws.merge_cells('B2:D2')
    cell = ws['B2']
    cell.value = f'Филиал {branch}'
    cell.font = Font(name='Arial', bold=True, size=11)
    cell.fill = PatternFill('solid', fgColor='FFFF00')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = [('A3','№'),('B3','Артикул'),('C3','Характеристика'),('D3','Кол-во')]
    for coord, val in headers:
        c = ws[coord]
        c.value = val
        c.font = Font(name='Arial', bold=True, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    ws.row_dimensions[3].height = 16
    for i, item in enumerate(items, 1):
        row = i + 3
        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=item.get('article', ''))
        ws.cell(row=row, column=3, value=item.get('size', '')).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4, value=item.get('qty', 1)).alignment = Alignment(horizontal='center')
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).font = Font(name='Arial', size=10)
        ws.row_dimensions[row].height = 15
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@app.route('/api/orders', methods=['POST'])
@login_required
def create_order():
    import json
    branch = session['branch'] if session['role'] == 'user' else request.form.get('branch', '').strip()
    responsible = request.form.get('responsible', '').strip()
    date = request.form.get('date', datetime.now().strftime('%Y-%m-%d'))
    priority = request.form.get('priority', 'Обычный')
    note = request.form.get('note', '').strip()
    items_json = request.form.get('items', '[]')
    if not branch or not responsible:
        return jsonify({'error': 'Заполните обязательные поля'}), 400
    items = json.loads(items_json)
    filename = None
    original_name = None
    f = request.files.get('file')
    if f and f.filename:
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ALLOWED:
            return jsonify({'error': 'Только .xlsx, .xls, .csv'}), 400
        original_name = f.filename
        filename = f'{datetime.now().strftime("%Y%m%d_%H%M%S")}_{secure_filename(f.filename)}'
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        f.save(file_path)
        upload_to_yadisk(file_path, original_name)
    if not filename and items:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        branch_slug = branch.replace(' ','_').replace('.','')
        original_name = f'ДОЗАКАЗ_{branch_slug}_{ts}.xlsx'
        filename = f'{ts}_{secure_filename(original_name)}'
        excel_buf = generate_order_excel(branch, items)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        with open(file_path, 'wb') as out:
            out.write(excel_buf.read())
        # Upload to Yandex Disk for permanent storage
        upload_to_yadisk(file_path, original_name)
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        'INSERT INTO orders (branch,responsible,date,priority,note,filename,original_name) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id',
        (branch, responsible, date, priority, note, filename, original_name)
    )
    order_id = cur.fetchone()['id']
    for item in items:
        cur.execute('INSERT INTO order_items (order_id,article,name,size,qty,wms_stock) VALUES (%s,%s,%s,%s,%s,%s)',
                   (order_id, item.get('article',''), item.get('name',''), item.get('size',''), item.get('qty',1), item.get('wms_stock',0)))
        # Deduct from WMS stock
        art = item.get('article','')
        size = item.get('size','')
        qty = item.get('qty',1)
        if art and size and qty:
            cur.execute('''
                UPDATE catalog SET wms_stock = GREATEST(0, wms_stock - %s)
                WHERE article=%s AND size=%s
            ''', (qty, art, size))
    conn.commit()
    cur.execute('SELECT * FROM orders WHERE id=%s', (order_id,))
    row = cur.fetchone()
    cur.close(); conn.close()
    return jsonify(dict(row)), 201

@app.route('/api/orders/<int:oid>/status', methods=['PATCH'])
@login_required
def update_status(oid):
    if session.get('role') not in ('admin', 'regional'):
        return jsonify({'error': 'Forbidden'}), 403
    data = request.get_json()
    status = data.get('status')
    if status not in ('Новая', 'В работе', 'Выполнена'):
        return jsonify({'error': 'Invalid status'}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute('UPDATE orders SET status=%s WHERE id=%s', (status, oid))
    conn.commit()
    cur.execute('SELECT * FROM orders WHERE id=%s', (oid,))
    row = cur.fetchone()
    cur.close(); conn.close()
    return jsonify(dict(row))

@app.route('/api/orders/<int:oid>/excel')
@login_required
def download_order_excel(oid):
    conn = get_db(); cur = conn.cursor()
    cur.execute('SELECT * FROM orders WHERE id=%s', (oid,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        return 'Not found', 404
    cur.execute('SELECT * FROM order_items WHERE order_id=%s', (oid,))
    items = cur.fetchall()
    cur.close(); conn.close()
    buf = generate_order_excel(row['branch'], [dict(i) for i in items])
    ts = row['date'].replace('-','') if row['date'] else datetime.now().strftime('%Y%m%d')
    fname = f"ДОЗАКАЗ_{row['branch'].replace(' ','_')}_{ts}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/download/<int:oid>')
@login_required
def download(oid):
    import urllib.request, urllib.parse, json as _json
    conn = get_db(); cur = conn.cursor()
    cur.execute('SELECT * FROM orders WHERE id=%s', (oid,))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row or not row['filename']:
        return 'Not found', 404
    local_path = os.path.join(app.config['UPLOAD_FOLDER'], row['filename'])
    # Try local first
    if os.path.exists(local_path):
        return send_from_directory(app.config['UPLOAD_FOLDER'], row['filename'],
                                   as_attachment=True, download_name=row['original_name'])
    # Try Yandex Disk
    try:
        yadisk_path = f'/li-ning-orders/{row["original_name"]}'
        url = ('https://cloud-api.yandex.net/v1/disk/resources/download?path='
               + urllib.parse.quote(yadisk_path, safe=''))
        req = urllib.request.Request(url, headers={'Authorization': f'OAuth {SYNC_TOKEN}'})
        with urllib.request.urlopen(req, timeout=10) as r:
            href = _json.loads(r.read()).get('href', '')
        if href:
            req2 = urllib.request.Request(href)
            with urllib.request.urlopen(req2, timeout=30) as r2:
                data = r2.read()
            # Save locally for next time
            with open(local_path, 'wb') as f:
                f.write(data)
            return send_from_directory(app.config['UPLOAD_FOLDER'], row['filename'],
                                       as_attachment=True, download_name=row['original_name'])
    except Exception as e:
        print(f'[DOWNLOAD] Yandex Disk error: {e}')
    return 'Файл не найден. Используйте кнопку Excel для скачивания.', 404

@app.route('/api/stats')
@login_required
def stats():
    conn = get_db(); cur = conn.cursor()
    if session['role'] == 'admin':
        cur.execute('SELECT COUNT(*) as c FROM orders'); total = cur.fetchone()['c']
        cur.execute("SELECT COUNT(*) as c FROM orders WHERE status='Новая'"); new = cur.fetchone()['c']
        cur.execute("SELECT COUNT(*) as c FROM orders WHERE status='В работе'"); inprog = cur.fetchone()['c']
        cur.execute("SELECT COUNT(*) as c FROM orders WHERE status='Выполнена'"); done = cur.fetchone()['c']
    elif session['role'] == 'regional':
        bs = session.get('branches', [])
        cur.execute('SELECT COUNT(*) as c FROM orders WHERE branch = ANY(%s)', (bs,)); total = cur.fetchone()['c']
        cur.execute("SELECT COUNT(*) as c FROM orders WHERE branch = ANY(%s) AND status='Новая'", (bs,)); new = cur.fetchone()['c']
        cur.execute("SELECT COUNT(*) as c FROM orders WHERE branch = ANY(%s) AND status='В работе'", (bs,)); inprog = cur.fetchone()['c']
        cur.execute("SELECT COUNT(*) as c FROM orders WHERE branch = ANY(%s) AND status='Выполнена'", (bs,)); done = cur.fetchone()['c']
    else:
        b = session['branch']
        cur.execute('SELECT COUNT(*) as c FROM orders WHERE branch=%s', (b,)); total = cur.fetchone()['c']
        cur.execute("SELECT COUNT(*) as c FROM orders WHERE branch=%s AND status='Новая'", (b,)); new = cur.fetchone()['c']
        cur.execute("SELECT COUNT(*) as c FROM orders WHERE branch=%s AND status='В работе'", (b,)); inprog = cur.fetchone()['c']
        cur.execute("SELECT COUNT(*) as c FROM orders WHERE branch=%s AND status='Выполнена'", (b,)); done = cur.fetchone()['c']
    cur.close(); conn.close()
    return jsonify({'total': total, 'new': new, 'in_progress': inprog, 'done': done})

@app.route('/api/catalog/upload', methods=['POST'])
@admin_required
def upload_catalog():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'Файл не найден'}), 400
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header_idx = None
        branch_cols = {}
        wms_col = None
        for i, row in enumerate(rows):
            row_vals = [str(c).strip() if c else '' for c in row]
            if 'Артикул' in row_vals:
                header_idx = i
                for j, val in enumerate(row_vals):
                    val_norm = val.replace('\u0421','C').replace('\u0441','c').replace('\u0415','E').replace('\u0435','e')
                    matched = None
                    for branch in BRANCHES:
                        branch_norm = branch.replace('\u0421','C').replace('\u0441','c').replace('\u0415','E').replace('\u0435','e')
                        if val_norm == branch_norm or val == branch:
                            matched = branch; break
                    if matched: branch_cols[matched] = j
                    if val == 'Склад WMS': wms_col = j
                break
        if header_idx is None:
            return jsonify({'error': 'Не найден заголовок таблицы'}), 400
        header_row = [str(c).strip() if c else '' for c in rows[header_idx]]
        name_col = size_col = season_col = category_col = None
        art_col = 0
        for j, val in enumerate(header_row):
            if val == 'Характеристика' and size_col is None: size_col = j
            if 'Номенклатура' in val and (',' in val or '.' in val) and 'Сезон' not in val and 'Вид' not in val and name_col is None: name_col = j
            if 'Сезон' in val: season_col = j
            if 'Вид' in val and category_col is None: category_col = j
        if name_col is None:
            for j, val in enumerate(header_row):
                if 'Номенклатура' in val and 'Сезон' not in val and 'Вид' not in val: name_col = j; break
        data_start = header_idx + 1
        for i in range(header_idx+1, min(header_idx+3, len(rows))):
            rv = [str(c).strip() if c else '' for c in rows[i]]
            if any('Доступно' in v for v in rv):
                data_start = i + 1; break
        conn = get_db(); cur = conn.cursor()
        # Save existing ABC/sold data
        cur.execute('SELECT article, MIN(abc) as abc, MAX(sold) as sold FROM catalog GROUP BY article')
        abc_data = {r['article']: (r['abc'], r['sold']) for r in cur.fetchall()}
        cur.execute('DELETE FROM catalog')
        cur.execute('DELETE FROM branch_stock')
        catalog_items = []
        branch_items = []
        for row in rows[data_start:]:
            if not row or not row[art_col]: continue
            art = str(row[art_col]).strip()
            if not art or art in ('None','nan'): continue
            size = ''
            if size_col is not None and row[size_col]:
                size = str(row[size_col]).strip()
            name = ''
            if name_col is not None and row[name_col]:
                name_full = str(row[name_col]).strip()
                if not size and ', ' in name_full:
                    parts = name_full.rsplit(', ', 1)
                    name_full = parts[0].strip()
                    size = parts[1].strip()
                art_base = art.rstrip('A')
                if name_full.startswith(art_base):
                    name_full = name_full[len(art_base):].strip()
                if size and name_full.endswith(f', {size}'):
                    name_full = name_full[:-len(f', {size}')].strip()
                name = name_full
            season = str(row[season_col]).strip() if season_col is not None and row[season_col] and str(row[season_col]) not in ('None','nan') else ''
            category = str(row[category_col]).strip() if category_col is not None and row[category_col] and str(row[category_col]) not in ('None','nan') else ''
            wms = 0
            if wms_col is not None and row[wms_col] and str(row[wms_col]) not in ('None','nan'):
                try: wms = int(float(str(row[wms_col])))
                except: pass
            catalog_items.append((art, name, size, wms, season, category))
            for branch, col in branch_cols.items():
                qty = row[col]
                if qty and str(qty) not in ('None','nan'):
                    try:
                        q = int(float(str(qty)))
                        if q > 0: branch_items.append((art, size, branch, q))
                    except: pass
        cur.executemany('INSERT INTO catalog (article,name,size,wms_stock,season,category) VALUES (%s,%s,%s,%s,%s,%s)', catalog_items)
        cur.executemany('INSERT INTO branch_stock (article,size,branch,qty) VALUES (%s,%s,%s,%s)', branch_items)
        # Restore ABC/sold data
        for art, (abc, sold) in abc_data.items():
            if abc and abc != 'C':
                cur.execute('UPDATE catalog SET abc=%s, sold=%s WHERE article=%s', (abc, sold, art))
        conn.commit(); cur.close(); conn.close()
        return jsonify({'ok': True, 'count': len(catalog_items)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/catalog', methods=['GET'])
@login_required
def get_catalog():
    search = request.args.get('search', '')
    page = int(request.args.get('page', 1))
    per_page = 50
    branch = session.get('branch') or request.args.get('branch', '')
    abc_filter = request.args.get('abc', '')
    season_filter = request.args.get('season', '')
    stock_filter = request.args.get('stock', '')
    conn = get_db(); cur = conn.cursor()

    # Step 1: Get paginated unique articles with aggregated data in ONE query
    where = '1=1'
    params = []
    if search:
        where += ' AND (article ILIKE %s OR name ILIKE %s)'
        params.extend([f'%{search}%', f'%{search}%'])
    if abc_filter:
        where += ' AND abc = %s'; params.append(abc_filter)
    if season_filter:
        where += ' AND season = %s'; params.append(season_filter)

    discount_filter = request.args.get('discount', '')
    if discount_filter == 'yes':
        where += ' AND discount > 0'

    cur.execute(f'''
        SELECT
            article,
            MIN(name) as name,
            MIN(abc) as abc,
            MAX(sold) as sold,
            MIN(season) as season,
            MIN(category) as category,
            SUM(wms_stock) as total_wms,
            MAX(discount) as discount
        FROM catalog
        WHERE {where}
        GROUP BY article
        HAVING 1=1
        {" AND SUM(wms_stock) > 0" if stock_filter == "yes" else ""}
        {" AND SUM(wms_stock) = 0" if stock_filter == "no" else ""}
        ORDER BY
            MAX(discount) DESC,
            CASE MIN(abc) WHEN 'A' THEN 1 WHEN 'B' THEN 2 ELSE 3 END,
            MAX(sold) DESC,
            article
        LIMIT {per_page} OFFSET {(page-1)*per_page}
    ''', params)
    articles = cur.fetchall()

    if not articles:
        cur.close(); conn.close()
        return jsonify([])

    art_list = [r['article'] for r in articles]

    # Step 2: Get all sizes for these articles in ONE query
    cur.execute('''
        SELECT article, size, wms_stock
        FROM catalog
        WHERE article = ANY(%s)
        ORDER BY article, size
    ''', (art_list,))
    all_sizes = {}
    for r in cur.fetchall():
        all_sizes.setdefault(r['article'], []).append({'size': r['size'], 'wms': r['wms_stock']})

    # Step 3: Get branch stock in ONE query if branch provided
    branch_stock_map = {}
    if branch:
        cur.execute('''
            SELECT article, size, qty
            FROM branch_stock
            WHERE article = ANY(%s) AND branch = %s
        ''', (art_list, branch))
        for r in cur.fetchall():
            branch_stock_map[(r['article'], r['size'])] = r['qty']

    # Get sales stats for all articles (7d, 30d, velocity)
    sales_stats = {}
    try:
        conn2 = get_db(); cur2 = conn2.cursor()
        art_list_for_sales = [art_row['article'] for art_row in articles]
        if art_list_for_sales:
            # Articles in sales are stored without 'A' suffix
            art_list_no_a = [a.rstrip('A') for a in art_list_for_sales]
            # Filter by branch if user role
            branch_filter = branch if branch else None
            if branch_filter:
                cur2.execute('''
                    SELECT article,
                        SUM(CASE WHEN sale_date >= CURRENT_DATE - 7 THEN qty ELSE 0 END) as sold_7d,
                        SUM(CASE WHEN sale_date >= CURRENT_DATE - 30 THEN qty ELSE 0 END) as sold_30d,
                        SUM(CASE WHEN sale_date >= CURRENT_DATE - 90 THEN qty ELSE 0 END) as sold_90d
                    FROM sales
                    WHERE article = ANY(%s) AND branch ILIKE %s
                    GROUP BY article
                ''', (art_list_no_a, branch_filter))
            else:
                cur2.execute('''
                    SELECT article,
                        SUM(CASE WHEN sale_date >= CURRENT_DATE - 7 THEN qty ELSE 0 END) as sold_7d,
                        SUM(CASE WHEN sale_date >= CURRENT_DATE - 30 THEN qty ELSE 0 END) as sold_30d,
                        SUM(CASE WHEN sale_date >= CURRENT_DATE - 90 THEN qty ELSE 0 END) as sold_90d
                    FROM sales
                    WHERE article = ANY(%s)
                    GROUP BY article
                ''', (art_list_no_a,))
            for r in cur2.fetchall():
                art_with_a = r['article'] + 'A'
                velocity = round(r['sold_30d'] / 30, 2) if r['sold_30d'] else 0
                stats = {
                    'sold_7d': r['sold_7d'] or 0,
                    'sold_30d': r['sold_30d'] or 0,
                    'sold_90d': r['sold_90d'] or 0,
                    'velocity': velocity
                }
                sales_stats[art_with_a] = stats
                sales_stats[r['article']] = stats
        cur2.close(); conn2.close()
    except Exception as e:
        import traceback
        print(f"[SALES STATS ERROR] {e}\n{traceback.format_exc()}")

    cur.close(); conn.close()

    result = []
    for art_row in articles:
        art = art_row['article']
        sizes = all_sizes.get(art, [])
        sizes_out = [{'size': s['size'], 'wms': s['wms'], 'branch': branch_stock_map.get((art, s['size']), 0)} for s in sizes]
        ss = sales_stats.get(art, {'sold_7d': 0, 'sold_30d': 0, 'sold_90d': 0, 'velocity': 0})
        result.append({
            'article': art,
            'name': art_row['name'] or '',
            'abc': art_row['abc'] or 'C',
            'sold': art_row['sold'] or 0,
            'season': art_row['season'] or '',
            'category': art_row['category'] or '',
            'total_wms': art_row['total_wms'] or 0,
            'discount': art_row['discount'] or 0,
            'sold_7d': ss['sold_7d'],
            'sold_30d': ss['sold_30d'],
            'sold_90d': ss['sold_90d'],
            'velocity': ss['velocity'],
            'sizes': sizes_out
        })
    return jsonify(result)

@app.route('/api/catalog/distribution/<article>')
@login_required
def catalog_distribution(article):
    '''Get stock and sales by branch for an article'''
    conn = get_db(); cur = conn.cursor()
    # Branch stock
    cur.execute('''
        SELECT branch, SUM(qty) as stock
        FROM branch_stock WHERE article=%s OR article=%s
        GROUP BY branch ORDER BY stock DESC
    ''', (article, article.rstrip('A')))
    stocks = {r['branch']: int(r['stock'] or 0) for r in cur.fetchall()}
    # Sales by branch (90 days)
    cur.execute('''
        SELECT branch, SUM(qty) as sold
        FROM sales WHERE (article=%s OR article=%s)
        AND sale_date >= CURRENT_DATE - INTERVAL '90 days'
        GROUP BY branch ORDER BY sold DESC
    ''', (article, article.rstrip('A')))
    sales = {r['branch']: int(r['sold'] or 0) for r in cur.fetchall()}
    cur.close(); conn.close()
    # Combine
    all_branches = set(list(stocks.keys()) + list(sales.keys()))
    result = []
    for b in BRANCHES:
        s = stocks.get(b, 0)
        sold = sales.get(b, 0)
        if s > 0 or sold > 0:
            result.append({'branch': b, 'stock': s, 'sold_90': sold})
    result.sort(key=lambda x: -x['stock'])
    return jsonify(result)

@app.route('/api/catalog/seasons')
@login_required
def get_seasons():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT DISTINCT season FROM catalog WHERE season != '' AND season IS NOT NULL ORDER BY season DESC")
    seasons = [r['season'] for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify(seasons)

@app.route('/api/catalog/abc', methods=['POST'])
@admin_required
def update_abc():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'Файл не найден'}), 400
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        conn = get_db(); cur = conn.cursor()
        count = 0
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            art_col = sold_col = None
            for i, row in enumerate(rows):
                row_s = [str(c).strip() if c else '' for c in row]
                if any(x in row_s for x in ['ITEM NO.','Одежда','Обувь']):
                    for j, v in enumerate(row_s):
                        if v in ('ITEM NO.','Одежда','Обувь'): art_col = j
                        if v == 'Продано': sold_col = j
                    for row2 in rows[i+1:]:
                        if not row2 or not row2[art_col]: continue
                        art = str(row2[art_col]).strip()
                        if art in ('Одежда','Обувь','Аксессуары','nan',''): continue
                        sold = int(float(str(row2[sold_col]))) if sold_col and row2[sold_col] else 0
                        abc = 'A' if sold >= 25 else 'B' if sold >= 15 else 'C'
                        cur.execute('UPDATE catalog SET abc=%s, sold=%s WHERE article LIKE %s', (abc, sold, art+'%'))
                        count += 1
                    break
        conn.commit(); cur.close(); conn.close()
        return jsonify({'ok': True, 'updated': count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/catalog/abc-upload', methods=['POST'])
@admin_required
def upload_abc_file():
    '''Upload ABC Excel file and sync catalog'''
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'Файл не найден'}), 400
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        # Try Export sheet first, then first sheet
        ws = wb['Export'] if 'Export' in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        # Find header row
        header_idx = None
        for i, row in enumerate(rows):
            if row and row[1] == 'Season3' or (row and str(row[0]) == 'DIVISION'):
                header_idx = i; break
        if header_idx is None:
            return jsonify({'error': 'Не найден заголовок'}), 400
        conn = get_db(); cur = conn.cursor()
        count = 0
        abc_map = {'A++': 'A', 'A+': 'A', 'A': 'A', 'A-': 'A', 'B': 'B', 'B+': 'B', 'B-': 'B', 'C': 'C', 'C+': 'C', 'C-': 'C'}
        for row in rows[header_idx+1:]:
            if not row or not row[2]: continue
            art = str(row[2]).strip()
            if not art or art in ('None', 'nan', ''): continue
            total_abc_raw = str(row[-1]).strip() if row[-1] else 'C'
            try: total_sold = int(float(str(row[-5]))) if row[-5] else 0
            except: total_sold = 0
            abc = abc_map.get(total_abc_raw, 'C')
            cur.execute('UPDATE catalog SET abc=%s, sold=%s WHERE article=%s OR article=%s',
                       (abc, total_sold, art, art+'A'))
            count += cur.rowcount
        conn.commit(); cur.close(); conn.close()
        return jsonify({'ok': True, 'updated': count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/catalog/abc-sync', methods=['POST'])
@admin_required
def abc_sync():
    data = request.get_json()
    updates = data.get('updates', [])
    if not updates:
        return jsonify({'error': 'Нет данных'}), 400
    conn = get_db(); cur = conn.cursor()
    count = 0
    for u in updates:
        art = u.get('article', '').strip()
        abc = u.get('abc', 'C')
        sold = u.get('sold', 0)
        # Match: exact, with A suffix, without A suffix
        cur.execute('UPDATE catalog SET abc=%s, sold=%s WHERE article=%s OR article=%s OR article=%s',
                   (abc, sold, art, art+'A', art.rstrip('A')))
        count += cur.rowcount
    conn.commit(); cur.close(); conn.close()
    return jsonify({'ok': True, 'updated': count})

@app.route('/api/products', methods=['GET'])
@login_required
def get_products():
    conn = get_db(); cur = conn.cursor()
    cur.execute('SELECT * FROM top_products ORDER BY sold DESC')
    rows = cur.fetchall()
    cur.close(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/products/upload', methods=['POST'])
@admin_required
def upload_products():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'Файл не найден'}), 400
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        products = []
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            art_col = sold_col = stock_col = cat_col = None
            header_row_idx = None
            for i, row in enumerate(rows):
                row_str = [str(c).strip() if c else '' for c in row]
                if any(x in row_str for x in ['ITEM NO.','Одежда','Обувь']):
                    header_row_idx = i
                    for j, cell in enumerate(row_str):
                        if cell in ('ITEM NO.','Одежда','Обувь','Аксессуары'): art_col = j
                        if cell == 'Продано': sold_col = j
                        if cell == 'Остаток': stock_col = j
                        if j == 1: cat_col = j
                    break
            if header_row_idx is not None and art_col is not None:
                current_cat = 'APP'
                for row in rows[header_row_idx+1:]:
                    if not row or not row[art_col]: continue
                    art = str(row[art_col]).strip()
                    if art in ('Одежда','Обувь','Аксессуары','nan',''):
                        if art == 'Обувь': current_cat = 'FTW'
                        elif art == 'Аксессуары': current_cat = 'ACC'
                        continue
                    cat = str(row[cat_col]).strip() if cat_col and row[cat_col] else current_cat
                    sold = int(float(str(row[sold_col]))) if sold_col and row[sold_col] and str(row[sold_col]) != 'None' else 0
                    stock = int(float(str(row[stock_col]))) if stock_col and row[stock_col] and str(row[stock_col]) != 'None' else 0
                    if art and art != 'None' and len(art) > 3:
                        products.append((art, cat, sold, stock))
        conn = get_db(); cur = conn.cursor()
        cur.execute('DELETE FROM top_products')
        cur.executemany('INSERT INTO top_products (article,category,sold,stock) VALUES (%s,%s,%s,%s)', products)
        conn.commit(); cur.close(); conn.close()
        return jsonify({'ok': True, 'count': len(products)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/analytics/orders')
@login_required
def analytics_orders():
    if session.get('role') not in ('admin', 'regional'):
        return jsonify({'error': 'Forbidden'}), 403
    conn = get_db(); cur = conn.cursor()
    if session['role'] == 'regional':
        bs = session.get('branches', [])
        cur.execute('''SELECT oi.article, SUM(oi.qty) as total_qty, COUNT(DISTINCT oi.order_id) as order_count
            FROM order_items oi JOIN orders o ON o.id = oi.order_id
            WHERE o.branch = ANY(%s) AND oi.article != %s AND oi.article != %s
            GROUP BY oi.article ORDER BY total_qty DESC LIMIT 20''', (bs, '', 'None'))
        top_articles = cur.fetchall()
        cur.execute('''SELECT o.branch, SUM(oi.qty) as total_qty, COUNT(DISTINCT o.id) as order_count
            FROM orders o JOIN order_items oi ON o.id = oi.order_id
            WHERE o.branch = ANY(%s)
            GROUP BY o.branch ORDER BY total_qty DESC''', (bs,))
        branch_totals = cur.fetchall()
    else:
        cur.execute('''SELECT article, SUM(qty) as total_qty, COUNT(DISTINCT order_id) as order_count
            FROM order_items WHERE article != '' AND article != 'None'
            GROUP BY article ORDER BY total_qty DESC LIMIT 20''')
        top_articles = cur.fetchall()
        cur.execute('''SELECT o.branch, SUM(oi.qty) as total_qty, COUNT(DISTINCT o.id) as order_count
            FROM orders o JOIN order_items oi ON o.id = oi.order_id
            GROUP BY o.branch ORDER BY total_qty DESC''')
        branch_totals = cur.fetchall()
    cur.close(); conn.close()
    return jsonify({'top_articles': [dict(r) for r in top_articles], 'branch_totals': [dict(r) for r in branch_totals]})

# ===== SALES =====

@app.route('/api/sales/upload', methods=['POST'])
@admin_required
def upload_sales():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'Файл не найден'}), 400
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header_idx = None
        for i, row in enumerate(rows):
            row_s = [str(c).strip() if c else '' for c in row]
            if 'Арт' in row_s or 'Артикул' in row_s:
                header_idx = i; break
        if header_idx is None:
            return jsonify({'error': 'Не найден заголовок'}), 400
        h1 = [str(c).strip() if c else '' for c in rows[header_idx]]
        season_col = art_col = cat_col = branch_col = ref_col = qty_col = price_col = amount_col = None
        for j, v in enumerate(h1):
            if 'Сезон' in v: season_col = j
            if 'Артикул' in v: art_col = j
            if 'Вид' in v and cat_col is None: cat_col = j
            if v == 'Магазин': branch_col = j
            if v == 'Ссылка': ref_col = j
            if v == 'Количество': qty_col = j
            if v == 'Цена': price_col = j
            if v == 'Сумма': amount_col = j
        replace = request.form.get('replace', 'false') == 'true'
        conn = get_db(); cur = conn.cursor()
        if replace:
            cur.execute('DELETE FROM sales')
        inserted = 0
        for row in rows[header_idx+1:]:
            if not row: continue
            art = str(row[art_col]).strip() if art_col is not None and row[art_col] else ''
            if not art or art in ('None','nan',''): continue
            season = str(row[season_col]).strip() if season_col is not None and row[season_col] else ''
            cat = str(row[cat_col]).strip() if cat_col is not None and row[cat_col] else ''
            branch = str(row[branch_col]).strip() if branch_col is not None and row[branch_col] else ''
            ref = str(row[ref_col]).strip() if ref_col is not None and row[ref_col] else ''
            try: qty = int(float(str(row[qty_col]).replace(' ','').replace(' ',''))) if qty_col is not None and row[qty_col] and str(row[qty_col]) not in ('None','nan') else 1
            except: qty = 1
            try: price = float(str(row[price_col]).replace(' ','').replace(' ','').replace(',','.')) if price_col is not None and row[price_col] and str(row[price_col]) not in ('None','nan') else 0
            except: price = 0
            try: amount = float(str(row[amount_col]).replace(' ','').replace(' ','').replace(',','.')) if amount_col is not None and row[amount_col] and str(row[amount_col]) not in ('None','nan') else 0
            except: amount = 0
            sale_date = None
            m = re.search(r'от (\d{2})\.(\d{2})\.(\d{4})', ref)
            if m:
                try: sale_date = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
                except: pass
            cur.execute(
                'INSERT INTO sales (season,article,category,branch,sale_date,qty,price,amount) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)',
                (season, art.rstrip('A'), cat, branch, sale_date, qty, price, amount)
            )
            inserted += 1
        conn.commit(); cur.close(); conn.close()
        return jsonify({'ok': True, 'inserted': inserted})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sales/analytics', methods=['GET'])
@login_required
def sales_analytics():
    period = request.args.get('period', 'month')
    category = request.args.get('category', '')
    branch = request.args.get('branch', '')
    limit = int(request.args.get('limit', 20))
    conn = get_db(); cur = conn.cursor()
    extra = ''
    params = []
    if period == 'week':
        extra += " AND sale_date >= CURRENT_DATE - INTERVAL '7 days'"
    elif period == 'month':
        extra += " AND sale_date >= CURRENT_DATE - INTERVAL '30 days'"
    elif period == 'season':
        extra += " AND sale_date >= CURRENT_DATE - INTERVAL '90 days'"
    if session['role'] == 'regional':
        bs = session.get('branches', [])
        if bs:
            extra += ' AND branch = ANY(%s)'; params.append(bs)
    elif branch:
        extra += ' AND branch = %s'; params.append(branch)
    if category:
        extra += ' AND category = %s'; params.append(category)
    cur.execute(f'''SELECT article, category,
        SUM(qty) as total_qty, SUM(amount) as total_amount, COUNT(*) as tx_count
        FROM sales WHERE 1=1 {extra}
        GROUP BY article, category ORDER BY total_qty DESC LIMIT %s
    ''', params + [limit])
    top_articles = cur.fetchall()
    cur.execute(f'''SELECT branch, SUM(qty) as total_qty, SUM(amount) as total_amount
        FROM sales WHERE 1=1 {extra} GROUP BY branch ORDER BY total_qty DESC
    ''', params)
    by_branch = cur.fetchall()
    cur.execute(f'''SELECT category, SUM(qty) as total_qty, SUM(amount) as total_amount
        FROM sales WHERE 1=1 {extra} GROUP BY category ORDER BY total_qty DESC
    ''', params)
    by_category = cur.fetchall()
    cur.execute(f'''SELECT SUM(qty) as total_qty, SUM(amount) as total_amount, COUNT(*) as tx_count
        FROM sales WHERE 1=1 {extra}
    ''', params)
    totals = cur.fetchone()
    cur.execute('SELECT MIN(sale_date) as min_date, MAX(sale_date) as max_date FROM sales')
    dates = cur.fetchone()
    cur.close(); conn.close()
    return jsonify({
        'top_articles': [dict(r) for r in top_articles],
        'by_branch': [dict(r) for r in by_branch],
        'by_category': [dict(r) for r in by_category],
        'totals': dict(totals) if totals else {},
        'dates': {'min': str(dates['min_date']) if dates and dates['min_date'] else None,
                  'max': str(dates['max_date']) if dates and dates['max_date'] else None}
    })

@app.route('/api/sales/info')
@admin_required
def sales_info():
    conn = get_db(); cur = conn.cursor()
    cur.execute('SELECT COUNT(*) as c FROM sales'); total = cur.fetchone()['c']
    cur.execute('SELECT MIN(sale_date) as mn, MAX(sale_date) as mx FROM sales'); dates = cur.fetchone()
    cur.close(); conn.close()
    return jsonify({'total': total, 'min_date': str(dates['mn']) if dates['mn'] else None, 'max_date': str(dates['mx']) if dates['mx'] else None})

# ===== PHOTOS =====

@app.route('/api/photos/<article>')
def get_photos(article):
    import urllib.request, urllib.parse, json as _json
    art_base = article.rstrip('A')
    if art_base in _photo_cache:
        return jsonify(_photo_cache[art_base])
    folder_path = f"{YADISK_FOLDER}/{art_base}"
    url = ("https://cloud-api.yandex.net/v1/disk/public/resources?"
        + urllib.parse.urlencode({'public_key': YADISK_PUBLIC_KEY, 'path': folder_path, 'limit': 20, 'preview_size': '400x400'}))
    try:
        req = urllib.request.Request(url, headers={'Authorization': f'OAuth {YADISK_TOKEN}'})
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = _json.loads(resp.read())
        items = data.get('_embedded', {}).get('items', [])
        photos = []
        for item in items:
            name = item.get('name', '')
            if name.lower().endswith(('.jpg','.jpeg','.png','.webp')):
                dl_url = ("https://cloud-api.yandex.net/v1/disk/public/resources/download?"
                    + urllib.parse.urlencode({'public_key': YADISK_PUBLIC_KEY, 'path': f"{folder_path}/{name}"}))
                req2 = urllib.request.Request(dl_url, headers={'Authorization': f'OAuth {YADISK_TOKEN}'})
                with urllib.request.urlopen(req2, timeout=8) as r2:
                    dl_data = _json.loads(r2.read())
                href = dl_data.get('href', '')
                if href:
                    proxy_key = f"{art_base}/{name}"
                    _photo_url_cache[proxy_key] = href
                    photos.append(f"/api/photo-proxy/{urllib.parse.quote(art_base)}/{urllib.parse.quote(name)}")
        result = {'photos': photos}
        _photo_cache[art_base] = result
        return jsonify(result)
    except Exception as e:
        return jsonify({'photos': [], 'error': str(e)})

@app.route('/api/photo-proxy/<art_base>/<filename>')
def photo_proxy(art_base, filename):
    import urllib.request, urllib.parse, json as _json
    key = f"{art_base}/{filename}"
    href = _photo_url_cache.get(key)
    if not href:
        folder_path = f"{YADISK_FOLDER}/{art_base}"
        dl_url = ("https://cloud-api.yandex.net/v1/disk/public/resources/download?"
            + urllib.parse.urlencode({'public_key': YADISK_PUBLIC_KEY, 'path': f"{folder_path}/{filename}"}))
        try:
            req = urllib.request.Request(dl_url, headers={'Authorization': f'OAuth {YADISK_TOKEN}'})
            with urllib.request.urlopen(req, timeout=8) as r:
                href = _json.loads(r.read()).get('href', '')
            _photo_url_cache[key] = href
        except:
            return 'Not found', 404
    try:
        req = urllib.request.Request(href)
        with urllib.request.urlopen(req, timeout=10) as r:
            data = r.read()
            ct = r.headers.get('Content-Type', 'image/jpeg')
        resp = Response(data, mimetype=ct)
        resp.headers['Cache-Control'] = 'public, max-age=86400'
        return resp
    except:
        return 'Error', 500

# ===== RECOMMENDATIONS =====

@app.route('/api/recommendations')
@login_required
def get_recommendations():
    branch = session.get('branch') or request.args.get('branch', '')
    if not branch:
        return jsonify([])
    conn = get_db(); cur = conn.cursor()

    # Get all A/B articles with their full size grid from WMS
    cur.execute('''
        SELECT c.article, MIN(c.name) as name, MIN(c.abc) as abc,
               MAX(c.sold) as sold, MIN(c.season) as season, MIN(c.category) as category,
               array_agg(c.size ORDER BY c.size) as all_sizes,
               array_agg(c.wms_stock ORDER BY c.size) as wms_stocks,
               SUM(c.wms_stock) as total_wms
        FROM catalog c
        WHERE c.abc IN ('A', 'B', 'C')
        GROUP BY c.article
        HAVING COUNT(c.size) >= 3
        ORDER BY MIN(c.abc), MAX(c.sold) DESC
        LIMIT 200
    ''')
    articles = cur.fetchall()

    # Get branch stock for all these articles
    art_list = [r['article'] for r in articles]
    if not art_list:
        cur.close(); conn.close()
        return jsonify([])

    cur.execute('''
        SELECT article, size, qty
        FROM branch_stock
        WHERE article = ANY(%s) AND branch = %s AND qty > 0
    ''', (art_list, branch))
    branch_stock = {}
    for r in cur.fetchall():
        branch_stock.setdefault(r['article'], {})[r['size']] = r['qty']

    cur.close(); conn.close()

    result = []
    for art_row in articles:
        art = art_row['article']
        all_sizes = art_row['all_sizes'] or []
        wms_stocks = art_row['wms_stocks'] or []
        branch_sizes = branch_stock.get(art, {})

        # Full size grid count
        full_grid = len(all_sizes)
        branch_has = len(branch_sizes)

        # Missing sizes (in full grid but not at branch)
        missing = []
        for i, size in enumerate(all_sizes):
            if size not in branch_sizes or branch_sizes[size] == 0:
                wms = wms_stocks[i] if i < len(wms_stocks) else 0
                missing.append({'size': size, 'wms': wms, 'available': wms > 0})

        # Show if any missing sizes available on WMS
        if not missing:
            continue

        result.append({
            'article': art,
            'name': art_row['name'] or '',
            'abc': art_row['abc'] or 'C',
            'sold': art_row['sold'] or 0,
            'season': art_row['season'] or '',
            'category': art_row['category'] or '',
            'total_wms': art_row['total_wms'] or 0,
            'full_grid': full_grid,
            'branch_has': branch_has,
            'missing_sizes': missing,
            'available_missing': [m for m in missing if m['available']]
        })

    # Sort: most missing available sizes first
    result.sort(key=lambda x: (-len(x['available_missing']), x['abc'], -x['sold']))
    return jsonify(result[:100])

@app.route('/api/recommendations/excel')
@login_required
def recommendations_excel():
    branch = session.get('branch') or request.args.get('branch', '')
    if not branch:
        return jsonify({'error': 'Филиал не указан'}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute('''
        SELECT c.article, MIN(c.name) as name, MIN(c.abc) as abc,
               MAX(c.sold) as sold, MIN(c.season) as season, MIN(c.category) as category,
               array_agg(c.size ORDER BY c.size) as all_sizes,
               array_agg(c.wms_stock ORDER BY c.size) as wms_stocks
        FROM catalog c
        WHERE c.abc IN ('A', 'B', 'C')
        GROUP BY c.article
        HAVING COUNT(c.size) >= 3
        ORDER BY MIN(c.abc), MAX(c.sold) DESC
        LIMIT 200
    ''')
    articles = cur.fetchall()
    art_list = [r['article'] for r in articles]
    cur.execute('SELECT article, size, qty FROM branch_stock WHERE article = ANY(%s) AND branch = %s AND qty > 0', (art_list, branch))
    branch_stock = {}
    for r in cur.fetchall():
        branch_stock.setdefault(r['article'], {})[r['size']] = r['qty']
    cur.close(); conn.close()

    # Build recommendations
    recs = []
    for art_row in articles:
        art = art_row['article']
        all_sizes = art_row['all_sizes'] or []
        wms_stocks = art_row['wms_stocks'] or []
        branch_sizes = branch_stock.get(art, {})
        missing = []
        for i, size in enumerate(all_sizes):
            if size not in branch_sizes or branch_sizes[size] == 0:
                wms = wms_stocks[i] if i < len(wms_stocks) else 0
                if wms > 0:
                    missing.append({'size': size, 'wms': wms})
        if missing:
            for m in missing:
                recs.append({
                    'article': art,
                    'name': art_row['name'] or '',
                    'abc': art_row['abc'] or 'C',
                    'season': art_row['season'] or '',
                    'category': art_row['category'] or '',
                    'size': m['size'],
                    'wms': m['wms'],
                    'qty': 1
                })

    # Generate Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Рекомендации'
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='E31837')
    # Title
    ws.merge_cells('A1:H1')
    title = ws['A1']
    title.value = f'Рекомендации к дозаказу — {branch}'
    title.font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    title.fill = PatternFill('solid', fgColor='E31837')
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24
    # Headers
    headers = ['№', 'Артикул', 'Название', 'Категория', 'Сезон', 'ABC', 'Размер', 'WMS склад', 'Заказать (шт)']
    col_widths = [4, 16, 35, 12, 10, 6, 8, 10, 12]
    for j, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[2].height = 18
    # ABC colors
    abc_colors = {'A': 'FFF3CD', 'B': 'D4EDDA', 'C': 'F8F9FA'}
    for i, rec in enumerate(recs, 1):
        row = i + 2
        vals = [i, rec['article'], rec['name'], rec['category'], rec['season'], rec['abc'], rec['size'], rec['wms'], 1]
        fill_color = abc_colors.get(rec['abc'], 'FFFFFF')
        for j, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=j, value=val)
            c.font = Font(name='Arial', size=10)
            c.border = border
            c.fill = PatternFill('solid', fgColor=fill_color)
            if j in (1, 6, 7, 8, 9):
                c.alignment = Alignment(horizontal='center')
        ws.row_dimensions[row].height = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    fname = f'Рекомендации_{branch.replace(" ","_")}_{ts}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/recommendations/cross')
@login_required
def cross_recommendations():
    '''Товары которые топ в других филиалах но отсутствуют в данном'''
    branch = session.get('branch') or request.args.get('branch', '')
    if not branch:
        return jsonify([])
    conn = get_db(); cur = conn.cursor()

    # Топ продаж по другим филиалам
    cur.execute('''
        SELECT s.article, MIN(c.name) as name, MIN(c.abc) as abc,
               SUM(s.qty) as total_sold, SUM(s.amount) as total_amount,
               MIN(c.season) as season, MIN(c.category) as category,
               array_agg(DISTINCT s.branch) as top_branches
        FROM sales s
        JOIN catalog c ON c.article = s.article OR c.article = s.article || 'A'
        WHERE s.branch != %s
        GROUP BY s.article
        HAVING SUM(s.qty) >= 5
        ORDER BY SUM(s.qty) DESC
        LIMIT 300
    ''', (branch,))
    top_elsewhere = cur.fetchall()

    if not top_elsewhere:
        cur.close(); conn.close()
        return jsonify([])

    art_list = [r['article'] for r in top_elsewhere]
    art_list_a = [a+'A' for a in art_list] + art_list

    # Продажи в данном филиале
    cur.execute('''
        SELECT article, SUM(qty) as sold
        FROM sales WHERE branch=%s AND article = ANY(%s)
        GROUP BY article
    ''', (branch, art_list + [a+'A' for a in art_list]))
    branch_sales = {r['article']: r['sold'] for r in cur.fetchall()}

    # Остатки в данном филиале
    cur.execute('''
        SELECT article, SUM(qty) as stock
        FROM branch_stock WHERE branch=%s AND article = ANY(%s)
        GROUP BY article
    ''', (branch, art_list_a))
    branch_stock = {r['article']: r['stock'] for r in cur.fetchall()}

    # WMS остатки
    cur.execute('''
        SELECT article, array_agg(size ORDER BY size) as sizes,
               array_agg(wms_stock ORDER BY size) as wms_stocks,
               SUM(wms_stock) as total_wms
        FROM catalog WHERE article = ANY(%s)
        GROUP BY article
    ''', (art_list_a,))
    wms_data = {r['article']: r for r in cur.fetchall()}

    cur.close(); conn.close()

    result = []
    for art_row in top_elsewhere:
        art = art_row['article']
        art_a = art + 'A'

        # Продажи в данном филиале
        my_sold = branch_sales.get(art, 0) + branch_sales.get(art_a, 0)
        # Остатки в данном филиале
        my_stock = branch_stock.get(art, 0) + branch_stock.get(art_a, 0)

        # Пропускаем если уже хорошо продаётся
        if my_sold >= art_row['total_sold'] * 0.3:
            continue

        # WMS данные
        wms = wms_data.get(art) or wms_data.get(art_a)
        total_wms = int(wms['total_wms']) if wms and wms['total_wms'] else 0
        sizes = []
        if wms:
            for s, w in zip(wms['sizes'] or [], wms['wms_stocks'] or []):
                if (w or 0) > 0:
                    sizes.append({'size': s, 'wms': int(w or 0)})

        if not sizes and total_wms == 0:
            continue

        # Топ филиалы где продаётся
        top_branches = [b for b in (art_row['top_branches'] or []) if b != branch][:3]

        result.append({
            'article': art,
            'name': art_row['name'] or '',
            'abc': art_row['abc'] or 'C',
            'season': art_row['season'] or '',
            'category': art_row['category'] or '',
            'total_sold_elsewhere': int(art_row['total_sold'] or 0),
            'my_sold': my_sold,
            'my_stock': my_stock,
            'total_wms': total_wms,
            'sizes': sizes,
            'top_branches': top_branches
        })

    # Сортировка: больший разрыв продаж → сначала
    result.sort(key=lambda x: -(x['total_sold_elsewhere'] - x['my_sold']))
    return jsonify(result[:80])

# ===== TRANSFERS =====

@app.route('/api/transfers', methods=['GET'])
@login_required
def get_transfers():
    conn = get_db(); cur = conn.cursor()
    role = session.get('role')
    branch = session.get('branch')
    if role == 'admin':
        cur.execute('SELECT * FROM transfers ORDER BY created_at DESC')
    elif branch in FLAGMANS:
        cur.execute('SELECT * FROM transfers WHERE to_branch=%s OR from_branch=%s ORDER BY created_at DESC', (branch, branch))
    else:
        cur.execute('SELECT * FROM transfers WHERE from_branch=%s ORDER BY created_at DESC', (branch,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/transfers', methods=['POST'])
@login_required
def create_transfer():
    # Only flagmans can create transfer requests
    branch = session.get('branch')
    if branch not in FLAGMANS and session.get('role') != 'admin':
        return jsonify({'error': 'Только флагманы могут создавать заявки на перемещение'}), 403
    data = request.get_json()
    from_branch = data.get('from_branch')
    to_branch = branch or data.get('to_branch')
    article = data.get('article')
    size = data.get('size')
    qty = data.get('qty', 1)
    note = data.get('note', '')
    if not all([from_branch, to_branch, article, size]):
        return jsonify({'error': 'Заполните все поля'}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute('INSERT INTO transfers (from_branch,to_branch,article,size,qty,note) VALUES (%s,%s,%s,%s,%s,%s) RETURNING id',
               (from_branch, to_branch, article, size, qty, note))
    tid = cur.fetchone()['id']
    conn.commit()
    cur.execute('SELECT * FROM transfers WHERE id=%s', (tid,))
    row = cur.fetchone()
    cur.close(); conn.close()
    return jsonify(dict(row)), 201

@app.route('/api/transfers/<int:tid>/status', methods=['PATCH'])
@admin_required
def update_transfer_status(tid):
    data = request.get_json()
    status = data.get('status')
    if status not in ('Новая', 'Подтверждена', 'Отклонена', 'Выполнена'):
        return jsonify({'error': 'Invalid status'}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute('UPDATE transfers SET status=%s WHERE id=%s', (status, tid))
    conn.commit()
    cur.execute('SELECT * FROM transfers WHERE id=%s', (tid,))
    row = cur.fetchone()
    cur.close(); conn.close()
    return jsonify(dict(row))

@app.route('/api/transfers/suggestions', methods=['GET'])
@login_required
def transfer_suggestions():
    '''Smart suggestions: A-items that sell well at flagman but poorly at donor branches'''
    branch = session.get('branch')
    if branch not in FLAGMANS and session.get('role') != 'admin':
        return jsonify([])
    target_branch = request.args.get('branch', branch)
    conn = get_db(); cur = conn.cursor()

    # Get A articles with their sales at target branch
    cur.execute('''
        SELECT c.article, MIN(c.name) as name, MIN(c.abc) as abc,
               MAX(c.sold) as total_sold,
               COALESCE(SUM(CASE WHEN bs.branch=%s THEN bs.qty ELSE 0 END), 0) as my_stock
        FROM catalog c
        LEFT JOIN branch_stock bs ON bs.article=c.article
        WHERE c.abc='A'
        GROUP BY c.article
        ORDER BY MAX(c.sold) DESC
        LIMIT 60
    ''', (target_branch,))
    a_articles = cur.fetchall()

    art_list = [r['article'] for r in a_articles]

    # Sales at target branch
    cur.execute('''
        SELECT article, SUM(qty) as sold FROM sales
        WHERE branch=%s AND article = ANY(%s)
        GROUP BY article
    ''', (target_branch, art_list))
    my_sales = {r['article']: r['sold'] for r in cur.fetchall()}

    # Sales + stock at ALL other branches
    cur.execute('''
        SELECT bs.article, bs.branch, bs.size, bs.qty as stock,
               COALESCE(s.branch_sold, 0) as branch_sold
        FROM branch_stock bs
        LEFT JOIN (
            SELECT article, branch, SUM(qty) as branch_sold
            FROM sales WHERE article = ANY(%s)
            GROUP BY article, branch
        ) s ON s.article=bs.article AND s.branch=bs.branch
        WHERE bs.article = ANY(%s) AND bs.branch != %s AND bs.qty > 0
        ORDER BY bs.article, branch_sold ASC, bs.qty DESC
    ''', (art_list, art_list, target_branch))
    donor_rows = cur.fetchall()

    # Group donors by article
    donors_by_art = {}
    for r in donor_rows:
        donors_by_art.setdefault(r['article'], []).append(dict(r))

    cur.close(); conn.close()

    result = []
    for art_row in a_articles:
        art = art_row['article']
        donors = donors_by_art.get(art, [])
        if not donors:
            continue

        my_sold = my_sales.get(art, 0)
        my_stock = int(art_row['my_stock'] or 0)

        # Only suggest if target branch sells well (or has demand) but low stock
        # OR if donors have stock but poor sales
        weak_donors = [d for d in donors if d['branch_sold'] < (my_sold * 0.5 + 1)]
        if not weak_donors:
            weak_donors = donors[:3]  # fallback: show worst performers

        result.append({
            'article': art,
            'name': art_row['name'] or '',
            'abc': art_row['abc'] or 'A',
            'total_sold': int(art_row['total_sold'] or 0),
            'my_sold': my_sold,
            'my_stock': my_stock,
            'donors': weak_donors[:5]  # top 5 weak donors
        })

    # Sort: biggest gap between my sales and donor sales first
    result.sort(key=lambda x: -(x['my_sold'] - min((d['branch_sold'] for d in x['donors']), default=0)))
    return jsonify(result[:30])

@app.route('/api/transfers/stats')
@login_required
def transfer_stats():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT COUNT(*) as c FROM transfers WHERE status='Новая'")
    new = cur.fetchone()['c']
    cur.close(); conn.close()
    return jsonify({'new': new})

# ===== SCHLOPKA =====

BRANCH_SHEET_MAP = {
    'алай': 'ALAYSKIY', 'атлас': 'ATLAS CHIMGAN', 'еко ': 'ECO PARK',
    'хай': 'HIGH TOWN PLAZA', 'медж': 'MAGIC CITY', 'малика': 'MALIKA',
    'новза': 'NOVZA', 'сккопус': 'Scopus Mall', 'шота': 'Shota Rustavely',
    'сити': 'TASHKENT CITY MALL', 'галерея': 'Yunusabad gallery'
}

@app.route('/api/schlopka', methods=['GET'])
@login_required
def get_schlopka_sessions():
    conn = get_db(); cur = conn.cursor()
    cur.execute('''
        SELECT s.*,
            COUNT(CASE WHEN i.status='Собран' OR i.status='Забрал' THEN 1 END) as collected,
            COUNT(i.id) as total_items
        FROM schlopka_sessions s
        LEFT JOIN schlopka_items i ON i.session_id = s.id
        GROUP BY s.id
        ORDER BY s.created_at DESC LIMIT 20
    ''')
    rows = cur.fetchall()
    cur.close(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/schlopka/upload', methods=['POST'])
@login_required
def upload_schlopka():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'Файл не найден'}), 400
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        items = []
        for sheet_name in wb.sheetnames:
            from_branch = sheet_name.strip()
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            # Find header row with 'Артикул'
            header_idx = None
            header_row = None
            for i, row in enumerate(rows[:6]):
                rv = [str(c).strip() if c else '' for c in row]
                if 'Артикул' in rv:
                    header_idx = i
                    header_row = rv
                    break
            if header_idx is None:
                continue
            # Detect columns from header
            art_col = name_col = cat_col = season_col = dest_col = None
            for j, v in enumerate(header_row):
                if v == 'Артикул': art_col = j
                if 'Характеристика' in v or ('Номенклатура' in v and ',' in v): name_col = j
                if 'Вид' in v and cat_col is None: cat_col = j
                if 'Сезон' in v and season_col is None: season_col = j
            # Last non-standard column = destination (куда)
            for j in range(len(header_row)-1, -1, -1):
                v = header_row[j]
                if v and v != 'Артикул' and 'Вид' not in v and 'Характеристика' not in v and 'Номенклатура' not in v and 'Сезон' not in v:
                    dest_col = j
                    break
            if art_col is None or dest_col is None:
                continue
            for row in rows[header_idx+1:]:
                if not row or not row[art_col]: continue
                art = str(row[art_col]).strip()
                if not art or art in ('None','nan',''): continue
                name_full = str(row[name_col]).strip() if name_col is not None and row[name_col] else ''
                category = str(row[cat_col]).strip() if cat_col is not None and row[cat_col] else ''
                to_dest = str(row[dest_col]).strip() if row[dest_col] else ''
                if not to_dest or to_dest in ('None','nan',''): continue
                # Extract size from name (last part after ', ')
                size = ''
                if ', ' in name_full:
                    parts = name_full.rsplit(', ', 1)
                    size = parts[1].strip()
                    name_full = parts[0].strip()
                # Clean article prefix
                art_base = art.rstrip('A')
                if name_full.startswith(art_base):
                    name_full = name_full[len(art_base):].strip()
                items.append((art, name_full or category, size, from_branch, to_dest))

        if not items:
            return jsonify({'error': 'Нет данных в файле'}), 400

        conn = get_db(); cur = conn.cursor()
        fname = secure_filename(f.filename) if f.filename else 'schlopka.xlsx'
        cur.execute('INSERT INTO schlopka_sessions (name,filename,created_by) VALUES (%s,%s,%s) RETURNING id',
                   (f.filename or 'Схлопка', fname, session.get('username')))
        sid = cur.fetchone()['id']
        # Store from_branch in branch col, to_branch in note col
        cur.executemany(
            'INSERT INTO schlopka_items (session_id,article,name,size,branch,qty,note) VALUES (%s,%s,%s,%s,%s,%s,%s)',
            [(sid, art, name, size, from_branch, 1, to_branch) for art, name, size, from_branch, to_branch in items]
        )
        conn.commit(); cur.close(); conn.close()
        return jsonify({'ok': True, 'session_id': sid, 'count': len(items)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/schlopka/<int:sid>', methods=['GET'])
@login_required
def get_schlopka_detail(sid):
    branch = session.get('branch')
    role = session.get('role')
    conn = get_db(); cur = conn.cursor()
    cur.execute('SELECT * FROM schlopka_sessions WHERE id=%s', (sid,))
    sess = cur.fetchone()
    if not sess:
        cur.close(); conn.close()
        return jsonify({'error': 'Не найдено'}), 404
    q = 'SELECT * FROM schlopka_items WHERE session_id=%s'
    params = [sid]
    if role == 'user' and branch:
        q += ' AND branch=%s'; params.append(branch)
    q += ' ORDER BY branch, article, size'
    cur.execute(q, params)
    items = cur.fetchall()
    # Branch summary for warehouse
    cur.execute('''
        SELECT branch,
            COUNT(*) as total,
            SUM(CASE WHEN branch_ready THEN 1 ELSE 0 END) as ready_count,
            SUM(CASE WHEN branch_taken THEN 1 ELSE 0 END) as taken_count,
            BOOL_AND(branch_ready) as all_ready,
            BOOL_AND(branch_taken) as all_taken
        FROM schlopka_items WHERE session_id=%s
        GROUP BY branch ORDER BY branch
    ''', (sid,))
    branches = cur.fetchall()
    cur.close(); conn.close()
    return jsonify({'session': dict(sess), 'items': [dict(i) for i in items], 'branches': [dict(b) for b in branches]})

@app.route('/api/schlopka/<int:sid>/bulk-status', methods=['PATCH'])
@login_required
def bulk_schlopka_status(sid):
    data = request.get_json()
    status = data.get('status')
    branch = data.get('branch', '')
    if status not in ('Не собран', 'В работе', 'Собран', 'Забрал'):
        return jsonify({'error': 'Invalid status'}), 400
    conn = get_db(); cur = conn.cursor()
    if branch:
        cur.execute('UPDATE schlopka_items SET status=%s, updated_at=NOW() WHERE session_id=%s AND branch=%s', (status, sid, branch))
    else:
        cur.execute('UPDATE schlopka_items SET status=%s, updated_at=NOW() WHERE session_id=%s', (status, sid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({'ok': True})

@app.route('/api/schlopka/items/<int:item_id>/status', methods=['PATCH'])
@login_required
def update_schlopka_status(item_id):
    data = request.get_json()
    status = data.get('status')
    if status not in ('Не собран', 'В работе', 'Собран', 'Забрал'):
        return jsonify({'error': 'Invalid status'}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute('UPDATE schlopka_items SET status=%s, updated_at=NOW() WHERE id=%s', (status, item_id))
    conn.commit()
    cur.execute('SELECT * FROM schlopka_items WHERE id=%s', (item_id,))
    row = cur.fetchone()
    cur.close(); conn.close()
    return jsonify(dict(row))

@app.route('/api/schlopka/<int:sid>/branch-ready', methods=['POST'])
@login_required
def branch_ready(sid):
    '''Branch marks their items as ready for pickup'''
    data = request.get_json() or {}
    branch = data.get('branch') or session.get('branch')
    if not branch:
        return jsonify({'error': 'Branch required'}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute('UPDATE schlopka_items SET branch_ready=TRUE WHERE session_id=%s AND branch=%s', (sid, branch))
    conn.commit(); cur.close(); conn.close()
    return jsonify({'ok': True})

@app.route('/api/schlopka/<int:sid>/branch-taken', methods=['POST'])
@login_required
def branch_taken(sid):
    '''Warehouse marks branch items as taken'''
    data = request.get_json() or {}
    branch = data.get('branch')
    if not branch:
        return jsonify({'error': 'Branch required'}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE schlopka_items SET branch_taken=TRUE, status='Забрал' WHERE session_id=%s AND branch=%s", (sid, branch))
    conn.commit(); cur.close(); conn.close()
    return jsonify({'ok': True})

@app.route('/api/schlopka/<int:sid>/notify', methods=['POST'])
@login_required
def notify_schlopka(sid):
    conn = get_db(); cur = conn.cursor()
    cur.execute('UPDATE schlopka_sessions SET ready_for_pickup=TRUE, ready_at=NOW() WHERE id=%s', (sid,))
    conn.commit(); cur.close(); conn.close()
    return jsonify({'ok': True})

@app.route('/api/schlopka/<int:sid>/notify', methods=['DELETE'])
@login_required
def cancel_notify_schlopka(sid):
    conn = get_db(); cur = conn.cursor()
    cur.execute('UPDATE schlopka_sessions SET ready_for_pickup=FALSE, ready_at=NULL WHERE id=%s', (sid,))
    conn.commit(); cur.close(); conn.close()
    return jsonify({'ok': True})

@app.route('/api/schlopka/<int:sid>', methods=['DELETE'])
@admin_required
def delete_schlopka(sid):
    conn = get_db(); cur = conn.cursor()
    cur.execute('DELETE FROM schlopka_sessions WHERE id=%s', (sid,))
    conn.commit(); cur.close(); conn.close()
    return jsonify({'ok': True})

def sync_catalog_from_yadisk():
    import urllib.request, urllib.parse, json as _json
    global _last_sync
    try:
        # Get download link
        full_path = SYNC_FOLDER + '/' + SYNC_FILENAME
        url = ('https://cloud-api.yandex.net/v1/disk/resources/download?path='
               + urllib.parse.quote(full_path, safe=''))
        req = urllib.request.Request(url, headers={'Authorization': f'OAuth {SYNC_TOKEN}'})
        with urllib.request.urlopen(req, timeout=15) as r:
            href = _json.loads(r.read()).get('href', '')
        if not href:
            print('[SYNC] No download link')
            return False
        # Download file
        req2 = urllib.request.Request(href)
        with urllib.request.urlopen(req2, timeout=60) as r2:
            file_data = r2.read()
        # Process file — reuse upload_catalog logic
        with app.app_context():
            wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            header_idx = None
            branch_cols = {}
            wms_col = None
            for i, row in enumerate(rows):
                row_vals = [str(c).strip() if c else '' for c in row]
                if 'Артикул' in row_vals:
                    header_idx = i
                    for j, val in enumerate(row_vals):
                        val_norm = val.replace('С','C').replace('с','c').replace('Е','E').replace('е','e')
                        for branch in BRANCHES:
                            branch_norm = branch.replace('С','C').replace('с','c').replace('Е','E').replace('е','e')
                            if val_norm == branch_norm or val == branch:
                                branch_cols[branch] = j; break
                        if val == 'Склад WMS': wms_col = j
                    break
            if header_idx is None:
                print('[SYNC] Header not found'); return False
            header_row = [str(c).strip() if c else '' for c in rows[header_idx]]
            name_col = size_col = season_col = category_col = None
            art_col = 0
            for j, val in enumerate(header_row):
                if val == 'Характеристика' and size_col is None: size_col = j
                if 'Номенклатура' in val and (',' in val or '.' in val) and 'Сезон' not in val and 'Вид' not in val and name_col is None: name_col = j
                if 'Сезон' in val: season_col = j
                if 'Вид' in val and category_col is None: category_col = j
            if name_col is None:
                for j, val in enumerate(header_row):
                    if 'Номенклатура' in val and 'Сезон' not in val and 'Вид' not in val: name_col = j; break
            data_start = header_idx + 1
            for i in range(header_idx+1, min(header_idx+3, len(rows))):
                rv = [str(c).strip() if c else '' for c in rows[i]]
                if any('Доступно' in v for v in rv):
                    data_start = i + 1; break
            conn = get_db(); cur = conn.cursor()
            # Save existing ABC/sold data before delete
            cur.execute('SELECT article, MIN(abc) as abc, MAX(sold) as sold FROM catalog GROUP BY article')
            abc_data = {r['article']: (r['abc'], r['sold']) for r in cur.fetchall()}
            cur.execute('DELETE FROM catalog')
            cur.execute('DELETE FROM branch_stock')
            catalog_items = []
            branch_items = []
            for row in rows[data_start:]:
                if not row or not row[art_col]: continue
                art = str(row[art_col]).strip()
                if not art or art in ('None','nan'): continue
                size = ''
                if size_col is not None and row[size_col]: size = str(row[size_col]).strip()
                name = ''
                if name_col is not None and row[name_col]:
                    name_full = str(row[name_col]).strip()
                    if not size and ', ' in name_full:
                        parts = name_full.rsplit(', ', 1); name_full = parts[0].strip(); size = parts[1].strip()
                    art_base = art.rstrip('A')
                    if name_full.startswith(art_base): name_full = name_full[len(art_base):].strip()
                    if size and name_full.endswith(f', {size}'): name_full = name_full[:-len(f', {size}')].strip()
                    name = name_full
                season = str(row[season_col]).strip() if season_col is not None and row[season_col] and str(row[season_col]) not in ('None','nan') else ''
                category = str(row[category_col]).strip() if category_col is not None and row[category_col] and str(row[category_col]) not in ('None','nan') else ''
                wms = 0
                if wms_col is not None and row[wms_col] and str(row[wms_col]) not in ('None','nan'):
                    try: wms = int(float(str(row[wms_col])))
                    except: pass
                catalog_items.append((art, name, size, wms, season, category))
                for branch, col in branch_cols.items():
                    qty = row[col]
                    if qty and str(qty) not in ('None','nan'):
                        try:
                            q = int(float(str(qty)))
                            if q > 0: branch_items.append((art, size, branch, q))
                        except: pass
            cur.executemany('INSERT INTO catalog (article,name,size,wms_stock,season,category) VALUES (%s,%s,%s,%s,%s,%s)', catalog_items)
            cur.executemany('INSERT INTO branch_stock (article,size,branch,qty) VALUES (%s,%s,%s,%s)', branch_items)
            # Restore ABC/sold data
            for art, (abc, sold) in abc_data.items():
                if abc and abc != 'C':
                    cur.execute('UPDATE catalog SET abc=%s, sold=%s WHERE article=%s', (abc, sold, art))
            conn.commit(); cur.close(); conn.close()
            _last_sync = datetime.now().strftime('%Y-%m-%d %H:%M')
            print(f'[SYNC] Done: {len(catalog_items)} items at {_last_sync}')
            return True
    except Exception as e:
        print(f'[SYNC] Error: {e}')
        return False

def sync_scheduler():
    time.sleep(10)  # wait for app to start
    while True:
        sync_catalog_from_yadisk()
        time.sleep(SYNC_INTERVAL_HOURS * 3600)

# Start background sync thread
sync_thread = Thread(target=sync_scheduler, daemon=True)
sync_thread.start()

@app.route('/api/debug/counts')
def debug_counts():
    conn = get_db(); cur = conn.cursor()
    cur.execute('SELECT COUNT(*) as c FROM catalog'); cat = cur.fetchone()['c']
    cur.execute('SELECT COUNT(DISTINCT article) as c FROM catalog'); arts = cur.fetchone()['c']
    cur.execute('SELECT COUNT(*) as c FROM sales'); sales = cur.fetchone()['c']
    cur.execute('SELECT COUNT(*) as c FROM orders'); orders = cur.fetchone()['c']
    cur.close(); conn.close()
    return jsonify({'catalog_rows': cat, 'unique_articles': arts, 'sales_rows': sales, 'orders': orders})

@app.route('/api/sync/status')
@admin_required
def sync_status():
    return jsonify({'last_sync': _last_sync, 'interval_hours': SYNC_INTERVAL_HOURS, 'folder': SYNC_FOLDER})

@app.route('/api/sync/now', methods=['POST'])
@admin_required
def sync_now():
    ok = sync_catalog_from_yadisk()
    return jsonify({'ok': ok, 'last_sync': _last_sync})

@app.route('/api/catalog/discount-upload', methods=['POST'])
@admin_required
def upload_discount():
    '''Upload list of articles with 70% discount'''
    f = request.files.get('file')
    data = request.get_json()
    conn = get_db(); cur = conn.cursor()
    
    if f:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        art_col = 0
        for i, row in enumerate(rows[:3]):
            rv = [str(c).strip().lower() if c else '' for c in row]
            if any('артикул' in v or 'article' in v for v in rv):
                for j, v in enumerate(rv):
                    if 'артикул' in v or 'article' in v:
                        art_col = j; break
                rows = rows[i+1:]
                break
        articles = []
        for row in rows:
            if row and row[art_col]:
                art = str(row[art_col]).strip()
                if art and art not in ('None', 'nan', ''):
                    articles.append(art)
    elif data:
        articles = data.get('articles', [])
    else:
        return jsonify({'error': 'Нет данных'}), 400
    
    # Reset all discounts first
    cur.execute("UPDATE catalog SET discount=0 WHERE discount=70")
    # Set 70% for uploaded articles
    count = 0
    for art in articles:
        cur.execute("UPDATE catalog SET discount=70 WHERE article=%s OR article=%s OR article=%s",
                   (art, art+'A', art.rstrip('A')))
        count += cur.rowcount
    conn.commit(); cur.close(); conn.close()
    return jsonify({'ok': True, 'updated': count, 'articles': len(articles)})

@app.route('/api/catalog/discount-clear', methods=['POST'])
@admin_required
def clear_discount():
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE catalog SET discount=0")
    conn.commit(); cur.close(); conn.close()
    return jsonify({'ok': True})

@app.route('/api/catalog/dead-stock')
@login_required
def dead_stock():
    '''Items on WMS with no sales in last 90 days'''
    conn = get_db(); cur = conn.cursor()
    now = datetime.now()
    month = now.month
    if month in (3,4,5): season_code = 'Q1'
    elif month in (6,7,8): season_code = 'Q2'
    elif month in (9,10,11): season_code = 'Q3'
    else: season_code = 'Q4'

    cur.execute('''
        SELECT c.article, MIN(c.name) as name, MIN(c.abc) as abc,
               MIN(c.season) as season, MIN(c.category) as category,
               SUM(c.wms_stock) as total_wms,
               MAX(c.sold) as total_sold,
               COALESCE(MAX(s.last_sale), NULL) as last_sale,
               COALESCE(SUM(s.qty_90), 0) as qty_90days
        FROM catalog c
        LEFT JOIN (
            SELECT article,
                   MAX(sale_date) as last_sale,
                   SUM(CASE WHEN sale_date >= CURRENT_DATE - INTERVAL '90 days' THEN qty ELSE 0 END) as qty_90
            FROM sales GROUP BY article
        ) s ON s.article = c.article OR s.article = c.article || 'A' OR s.article || 'A' = c.article
        WHERE c.wms_stock > 0
        GROUP BY c.article
        HAVING SUM(c.wms_stock) > 0
        AND COALESCE(SUM(s.qty_90), 0) = 0
        ORDER BY SUM(c.wms_stock) DESC
        LIMIT 100
    ''')
    rows = cur.fetchall()
    cur.close(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/ai/movement-plan', methods=['GET'])
@login_required
def ai_movement_plan():
    '''Generate Excel movement plan based on sales vs stock analysis'''
    conn = get_db(); cur = conn.cursor()
    
    my_branch = session.get('branch')
    role = session.get('role')

    # Get sales by article+branch
    cur.execute('''
        SELECT s.article, s.branch, SUM(s.qty) as sold,
               COALESCE(bs.total_stock, 0) as stock
        FROM sales s
        LEFT JOIN (
            SELECT article, branch, SUM(qty) as total_stock
            FROM branch_stock GROUP BY article, branch
        ) bs ON bs.article=s.article AND bs.branch=s.branch
        WHERE s.sale_date >= CURRENT_DATE - INTERVAL '90 days'
        GROUP BY s.article, s.branch, bs.total_stock
        ORDER BY s.article, SUM(s.qty) DESC
    ''')
    sales_rows = cur.fetchall()
    
    # Get ABC A articles
    cur.execute("SELECT DISTINCT article FROM catalog WHERE abc='A'")
    a_articles = {r['article'] for r in cur.fetchall()}
    cur.close(); conn.close()

    # Build movement plan
    by_article = {}
    for r in sales_rows:
        art = r['article']
        if art not in by_article:
            by_article[art] = []
        by_article[art].append({'branch': r['branch'], 'sold': int(r['sold'] or 0), 'stock': int(r['stock'] or 0)})

    movements = []
    for art, branches in by_article.items():
        if len(branches) < 2: continue
        branches.sort(key=lambda x: x['sold'], reverse=True)
        top = branches[0]  # best seller
        for donor in branches[1:]:
            gap = top['sold'] - donor['sold']
            if gap >= 3 and donor['stock'] >= 1:
                qty = min(donor['stock'], max(1, gap // 3))
                movements.append({
                    'article': art,
                    'from_branch': donor['branch'],
                    'to_branch': top['branch'],
                    'donor_sold': donor['sold'],
                    'receiver_sold': top['sold'],
                    'donor_stock': donor['stock'],
                    'qty': qty,
                    'priority': 'A' if art in a_articles else 'B'
                })

    # Filter by branch if not admin
    if my_branch and role != 'admin':
        movements = [m for m in movements if m['from_branch'] == my_branch or m['to_branch'] == my_branch]

    movements.sort(key=lambda x: (x['priority'], -(x['receiver_sold'] - x['donor_sold'])))

    # Generate Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'План перемещений'
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Title
    ws.merge_cells('A1:H1')
    t = ws['A1']
    t.value = f'План перемещений товаров — {datetime.now().strftime("%d.%m.%Y")}'
    t.font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    t.fill = PatternFill('solid', fgColor='E31837')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    headers = ['№', 'Артикул', 'ABC', 'Откуда', 'Продаж (донор)', 'Куда', 'Продаж (получатель)', 'Остаток у донора', 'Кол-во для передачи']
    widths = [4, 16, 5, 20, 14, 20, 18, 16, 16]
    fills = {'A': 'FFF3CD', 'B': 'D4EDDA'}
    
    for j, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='E31837')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[2].height = 18

    for i, m in enumerate(movements[:200], 1):
        row = i + 2
        fill_color = fills.get(m['priority'], 'FFFFFF')
        vals = [i, m['article'], m['priority'], m['from_branch'], m['donor_sold'],
                m['to_branch'], m['receiver_sold'], m['donor_stock'], m['qty']]
        for j, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=j, value=val)
            c.font = Font(name='Arial', size=10)
            c.border = border
            c.fill = PatternFill('solid', fgColor=fill_color)
            if j in (1,3,5,7,8,9):
                c.alignment = Alignment(horizontal='center')
        ws.row_dimensions[row].height = 15

    # Summary sheet
    ws2 = wb.create_sheet('Сводка')
    ws2['A1'] = 'Итого перемещений:'; ws2['B1'] = len(movements)
    ws2['A2'] = 'Категория A:'; ws2['B2'] = sum(1 for m in movements if m['priority']=='A')
    ws2['A3'] = 'Уникальных артикулов:'; ws2['B3'] = len(set(m['article'] for m in movements))
    
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    branch_slug = (my_branch or 'все').replace(' ','_') if my_branch and role!='admin' else 'все_филиалы'
    fname = f'план_перемещений_{branch_slug}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/ai/analyze', methods=['POST'])
@login_required
def ai_analyze():
    import urllib.request, json as _json
    data = request.get_json()
    prompt = data.get('prompt', '')
    if not prompt:
        return jsonify({'error': 'No prompt'}), 400
    try:
        payload = _json.dumps({
            'model': 'claude-sonnet-4-6',
            'max_tokens': 1500,
            'messages': [{'role': 'user', 'content': prompt}]
        }).encode()
        api_key = os.environ.get('ANTHROPIC_API_KEY', '')
        if not api_key:
            return jsonify({'error': 'ANTHROPIC_API_KEY не настроен в Railway Variables'}), 500
        req = urllib.request.Request(
            'https://api.anthropic.com/v1/messages',
            data=payload,
            headers={
                'Content-Type': 'application/json',
                'x-api-key': api_key,
                'anthropic-version': '2023-06-01'
            },
            method='POST'
        )
        with urllib.request.urlopen(req, timeout=30) as r:
            result = _json.loads(r.read())
        text = result.get('content', [{}])[0].get('text', 'Нет ответа')
        return jsonify({'text': text})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/distribution/run', methods=['POST'])
@login_required
def run_distribution():
    if session.get('role') not in ('admin', 'warehouse'):
        return jsonify({'error': 'Нет доступа'}), 403
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'Файл не найден'}), 400
    try:
        wb_in = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb_in.active
        all_rows = list(ws.iter_rows(values_only=True))

        # Find header row with 'Артикул'
        header_idx = None
        for i, row in enumerate(all_rows[:5]):
            if row and str(row[0]).strip() == 'Артикул':
                header_idx = i
                break
        if header_idx is None:
            return jsonify({'error': 'Не найден заголовок "Артикул"'}), 400

        header_row = all_rows[header_idx]
        data_rows  = all_rows[header_idx + 1:]

        # Find store columns (col 3 onwards after Артикул, Товар, Количество)
        stores = []
        store_cols = []
        for j, val in enumerate(header_row):
            if j >= 3 and val and str(val).strip() not in ('', 'None'):
                stores.append(str(val).strip())
                store_cols.append(j)

        n_stores = len(stores)
        if n_stores == 0:
            return jsonify({'error': 'Не найдены колонки магазинов'}), 400

        # Priority order of stores
        PRIORITY_ORDER = [
            'TASHKENT CITY MALL',
            'ALAYSKIY',
            'ATLAS CHIMGAN',
            'Shota Rustavely',
            'MAGIC CITY',
            'HIGH TOWN PLAZA',
            'MALIKA',
            'ECO PARK',
            'NOVZA',
            'Scopus Mall',
            'Yunusabad gallery',
            'M. BARAKA',
            'Family park',
            'UZBEGIM ANDIJAN',
        ]

        FLAGMANS = {'TASHKENT CITY MALL', 'ALAYSKIY', 'ATLAS CHIMGAN', 'Shota Rustavely'}
        KIDS_STORES = {'TASHKENT CITY MALL', 'MAGIC CITY'}
        POPULAR_SIZES = {'S', 'M', 'L', 'XL', '7.5', '8', '8.5', '9', '9.5', '10', '7,5', '8,5', '9,5'}
        MIN_QTY = 4  # minimum per article per store

        # Build store index map for sorting
        def store_priority(s):
            s = (s or '').strip()
            for i, p in enumerate(PRIORITY_ORDER):
                if p.lower() == s.lower():
                    return i
            return 99

        # Sort stores by priority
        sorted_store_indices = sorted(range(n_stores), key=lambda i: store_priority(stores[i]))

        from collections import OrderedDict

        # Group rows by article
        art_rows = OrderedDict()
        for row in data_rows:
            if not row or not row[0]: continue
            art = str(row[0]).strip()
            if not art or art in ('None', 'nan', 'Итого'): continue
            nom_full = str(row[1]).strip() if row[1] else ''
            qty = int(float(str(row[2]))) if row[2] and str(row[2]) not in ('None','nan','') else 0
            if qty <= 0: continue
            if art not in art_rows:
                art_rows[art] = []
            art_rows[art].append({'nom': nom_full, 'qty': qty})

        results = []; total_qty_in = 0; total_alloc = 0

        for art, rows in art_rows.items():
            is_kids = art.startswith('Y') or art.startswith('y')
            total_qty_in += sum(r['qty'] for r in rows)

            # Total qty for this article
            total_art_qty = sum(r['qty'] for r in rows)

            # Determine eligible stores
            # If total qty < MIN_QTY * n_all_stores → flagmans only
            # Kids (Y...) → only KIDS_STORES
            all_eligible = []
            flagman_eligible = []
            for si in sorted_store_indices:
                store_name = (stores[si] or '').strip()
                if is_kids and store_name not in KIDS_STORES:
                    continue
                all_eligible.append(si)
                if store_name in FLAGMANS:
                    flagman_eligible.append(si)

            # Decide: flagmans only or all?
            n_all = len(all_eligible)
            enough_for_all = total_art_qty >= MIN_QTY * n_all
            eligible = all_eligible if enough_for_all else flagman_eligible

            n_eligible = len(eligible)
            if n_eligible == 0:
                for r in rows:
                    results.append({'art':art,'nom':r['nom'],'qty':r['qty'],'is_kids':is_kids,
                                    'size':'','is_popular':False,'wms_left':r['qty'],'alloc':[0]*n_stores})
                continue

            # For each size: distribute round-robin (1 by 1) across eligible stores
            # Popular sizes (S/M/L/XL, 7.5-10) get double on second pass
            for r in rows:
                qty = r['qty']
                nom_stripped = r['nom'].strip()
                if nom_stripped.endswith(')') and '(' in nom_stripped:
                    size_hint = nom_stripped[nom_stripped.rfind('(')+1:-1].strip()
                elif ',' in nom_stripped:
                    size_hint = nom_stripped.rsplit(',', 1)[-1].strip()
                else:
                    size_hint = ''

                is_popular = size_hint in POPULAR_SIZES

                # Round-robin distribution: deal 1 card at a time to each store
                store_counts = {si: 0 for si in eligible}
                remaining = qty
                round_num = 0

                while remaining > 0:
                    gave_any = False
                    for si in eligible:
                        if remaining <= 0:
                            break
                        # Popular sizes: give 2 on first pass, 1 on subsequent
                        give = 2 if (is_popular and round_num == 0) else 1
                        give = min(give, remaining)
                        store_counts[si] += give
                        remaining -= give
                        gave_any = True
                    round_num += 1
                    if not gave_any:
                        break

                # Build alloc array
                alloc = [0] * n_stores
                for si in eligible:
                    alloc[si] = store_counts[si]

                # Check minimum: store must have >= MIN_QTY total for this article
                # We check this after distributing all sizes together
                # For now apply per-size distribution, min check done below
                total_alloc += sum(alloc)
                results.append({
                    'art': art,
                    'nom': r['nom'],
                    'qty': qty,
                    'is_kids': is_kids,
                    'size': size_hint,
                    'is_popular': is_popular,
                    'wms_left': remaining,
                    'alloc': alloc
                })

        # Post-process: zero out stores that got < MIN_QTY total for an article
        art_store_totals = {}  # (art, si) -> total
        for r in results:
            art = r['art']
            for si, give in enumerate(r['alloc']):
                key = (art, si)
                art_store_totals[key] = art_store_totals.get(key, 0) + give

        for r in results:
            art = r['art']
            for si in range(n_stores):
                if art_store_totals.get((art, si), 0) < MIN_QTY:
                    freed = r['alloc'][si]
                    r['alloc'][si] = 0
                    r['wms_left'] += freed
                    total_alloc -= freed
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = 'Распределение'
        from openpyxl.utils import get_column_letter as gcl
        thin=Side(style='thin',color='CCCCCC'); brd=Border(left=thin,right=thin,top=thin,bottom=thin)
        ctr=Alignment(horizontal='center',vertical='center',wrap_text=True); lft=Alignment(horizontal='left',vertical='center')
        hdr_fill=PatternFill('solid',fgColor='1A1A2E'); alloc_fill=PatternFill('solid',fgColor='C8E6C9')
        zero_fill=PatternFill('solid',fgColor='FFEBEE'); total_fill=PatternFill('solid',fgColor='E3F2FD')
        wms_fill=PatternFill('solid',fgColor='FFF3E0')

        fixed_hdrs=['Артикул','Номенклатура','Кол-во (приход)','Остаток']
        store_hdrs=[s for s in stores]
        all_hdrs=fixed_hdrs+store_hdrs

        for ci,h in enumerate(all_hdrs,1):
            cell=ws_out.cell(row=1,column=ci,value=h)
            cell.fill=hdr_fill; cell.font=Font(bold=True,color='FFFFFF',size=9)
            cell.alignment=ctr; cell.border=brd
        ws_out.row_dimensions[1].height=36

        kids_fill = PatternFill('solid',fgColor='E8F5E9')
        for ri,row in enumerate(results,2):
            vals=[row['art'],row['nom'],row['qty'],row['wms_left']]
            for ci,v in enumerate(vals,1):
                cell=ws_out.cell(row=ri,column=ci,value=v)
                cell.border=brd; cell.font=Font(size=9)
                cell.alignment=lft if ci==2 else ctr
                if ci==3: cell.fill=wms_fill
                if ci==4 and (v or 0)>0: cell.fill=PatternFill('solid',fgColor='FFF9C4')
                if row.get('is_kids'): cell.font=Font(size=9,color='1B5E20',bold=(ci==1))
            for i,give in enumerate(row['alloc']):
                ci=5+i
                cell=ws_out.cell(row=ri,column=ci,value=give if give else None)
                cell.border=brd; cell.alignment=ctr
                cell.font=Font(size=9,bold=(give>0),color=('1B5E20' if give>0 else '000000'))
                cell.fill=alloc_fill if give>0 else zero_fill

        last=len(results)+2
        ws_out.cell(row=last,column=1,value='ИТОГО').font=Font(bold=True,size=10)
        ws_out.cell(row=last,column=3,value=total_qty_in).font=Font(bold=True)
        ws_out.cell(row=last,column=3).fill=total_fill
        ws_out.cell(row=last,column=4,value=total_qty_in-total_alloc).font=Font(bold=True)
        ws_out.cell(row=last,column=4).fill=total_fill
        for i in range(n_stores):
            ci=5+i; t=sum(r['alloc'][i] for r in results)
            cell=ws_out.cell(row=last,column=ci,value=t)
            cell.font=Font(bold=True,size=10); cell.fill=total_fill
            cell.alignment=ctr; cell.border=brd
        for ci in range(1,5): ws_out.cell(row=last,column=ci).fill=total_fill

        widths=[14,40,10,9]+[12]*n_stores
        for i,w in enumerate(widths,1): ws_out.column_dimensions[gcl(i)].width=w
        ws_out.freeze_panes='E2'

        ws2=wb_out.create_sheet('Сводка')
        hdrs2=['Магазин','Получает (шт)','Артикулов']
        for ci,h in enumerate(hdrs2,1):
            cell=ws2.cell(row=1,column=ci,value=h)
            cell.fill=hdr_fill; cell.font=Font(bold=True,color='FFFFFF'); cell.alignment=ctr
        ws2.row_dimensions[1].height=24
        store_totals=[sum(r['alloc'][i] for r in results) for i in range(n_stores)]
        store_arts=[sum(1 for r in results if r['alloc'][i]>0) for i in range(n_stores)]
        for i,(store,tot,arts) in enumerate(zip(stores,store_totals,store_arts),2):
            ws2.cell(row=i,column=1,value=store).font=Font(size=10)
            ws2.cell(row=i,column=2,value=tot).alignment=ctr
            ws2.cell(row=i,column=3,value=arts).alignment=ctr
            if tot>0:
                ws2.cell(row=i,column=2).fill=alloc_fill
                ws2.cell(row=i,column=2).font=Font(bold=True,color='1B5E20')
        r_total=n_stores+2
        ws2.cell(row=r_total,column=1,value='ВСЕГО').font=Font(bold=True)
        ws2.cell(row=r_total,column=2,value=total_alloc).font=Font(bold=True)
        for col,w in zip(range(1,4),[28,14,12]): ws2.column_dimensions[gcl(col)].width=w
        output=io.BytesIO(); wb_out.save(output); output.seek(0)
        fname=f'distribution_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return send_file(output,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',as_attachment=True,download_name=fname)
    except Exception as e:
        import traceback
        return jsonify({'error':str(e),'tb':traceback.format_exc()}),500


# ===== ZERO SALES =====

@app.route('/api/zero-sales')
@login_required
def zero_sales():
    """
    Артикулы без продаж за 30 дней.
    Для каждого филиала показывает:
    - Остаток в филиале
    - Суммарный остаток по сети
    - Флаг битой сетки (< 3 размеров в филиале)
    - Флаг авто-схлопки (суммарный остаток < 50)
    """
    role = session.get('role')
    # Branch: from session (store user) or from URL param (admin selecting branch)
    branch = session.get('branch') or request.args.get('branch', '')
    days = int(request.args.get('days', 30))

    conn = get_db(); cur = conn.cursor()

    # Single branch query (user role OR admin with selected branch)
    if branch:
        cur.execute("""
            SELECT
                bs.article,
                MIN(c.name) as name,
                MIN(c.abc) as abc,
                MIN(c.season) as season,
                MIN(c.category) as category,
                bs.branch,
                SUM(bs.qty) as branch_stock,
                COUNT(DISTINCT bs.size) as branch_sizes,
                SUM(c.wms_stock) as wms_stock,
                array_agg(DISTINCT bs.size ORDER BY bs.size) as branch_size_list,
                array_agg(c.size ORDER BY c.size) as all_wms_sizes,
                array_agg(c.wms_stock ORDER BY c.size) as all_wms_stocks
            FROM branch_stock bs
            JOIN catalog c ON c.article = bs.article
            WHERE bs.branch = %s AND bs.qty > 0
            AND bs.article NOT IN (
                SELECT DISTINCT article FROM sales
                WHERE branch = %s
                AND sale_date >= CURRENT_DATE - %s
            )
            AND bs.article NOT IN (
                SELECT DISTINCT article || 'A' FROM sales
                WHERE branch = %s
                AND sale_date >= CURRENT_DATE - %s
            )
            AND c.discount = 0
            GROUP BY bs.article, bs.branch
            ORDER BY MIN(c.abc), SUM(bs.qty) DESC
            LIMIT 200
        """, (branch, branch, days, branch, days))
    else:
        # Admin: all branches
        cur.execute("""
            SELECT
                bs.article,
                MIN(c.name) as name,
                MIN(c.abc) as abc,
                MIN(c.season) as season,
                MIN(c.category) as category,
                bs.branch,
                SUM(bs.qty) as branch_stock,
                COUNT(DISTINCT bs.size) as branch_sizes,
                SUM(c.wms_stock) as wms_stock,
                array_agg(DISTINCT bs.size ORDER BY bs.size) as branch_size_list,
                array_agg(c.size ORDER BY c.size) as all_wms_sizes,
                array_agg(c.wms_stock ORDER BY c.size) as all_wms_stocks
            FROM branch_stock bs
            JOIN catalog c ON c.article = bs.article
            WHERE bs.qty > 0
            AND bs.article NOT IN (
                SELECT DISTINCT article FROM sales
                WHERE sale_date >= CURRENT_DATE - %s
            )
            AND bs.article NOT IN (
                SELECT DISTINCT article || 'A' FROM sales
                WHERE sale_date >= CURRENT_DATE - %s
            )
            AND c.discount = 0
            GROUP BY bs.article, bs.branch
            ORDER BY MIN(c.abc), SUM(bs.qty) DESC
            LIMIT 500
        """, (days, days))

    rows = cur.fetchall()

    # Суммарный остаток по всей сети для каждого артикула
    art_list = list(set(r['article'] for r in rows))
    network_stock = {}
    network_sizes = {}
    if art_list:
        cur.execute("""
            SELECT article, SUM(qty) as total, COUNT(DISTINCT size) as sizes
            FROM branch_stock WHERE article = ANY(%s)
            GROUP BY article
        """, (art_list,))
        for r in cur.fetchall():
            network_stock[r['article']] = int(r['total'] or 0)
            network_sizes[r['article']] = int(r['sizes'] or 0)

    cur.close(); conn.close()

    result = []
    seen = set()

    for r in rows:
        art = r['article']
        total_network = network_stock.get(art, 0)
        total_sizes = network_sizes.get(art, 0)
        branch_sizes_count = int(r['branch_sizes'] or 0)

        is_broken_grid = branch_sizes_count < 3
        needs_schlopka = total_network < 50 and total_network > 0

        # Build WMS sizes info: which sizes available on WMS but missing in branch
        branch_size_list = r['branch_size_list'] or []
        all_wms_sizes = r['all_wms_sizes'] or []
        all_wms_stocks = r['all_wms_stocks'] or []
        wms_sizes_detail = []
        for i, sz in enumerate(all_wms_sizes):
            wms_qty = int(all_wms_stocks[i] or 0) if i < len(all_wms_stocks) else 0
            has_in_branch = sz in branch_size_list
            wms_sizes_detail.append({
                'size': sz,
                'wms': wms_qty,
                'in_branch': has_in_branch,
                'can_order': wms_qty > 0 and not has_in_branch
            })
        total_wms = int(r['wms_stock'] or 0) if 'wms_stock' in r else 0
        can_order_count = sum(1 for s in wms_sizes_detail if s['can_order'])

        item = {
            'article': art,
            'name': r['name'] or '',
            'abc': r['abc'] or 'C',
            'season': r['season'] or '',
            'category': r['category'] or '',
            'branch_stock': int(r['branch_stock'] or 0),
            'branch_sizes': branch_sizes_count,
            'network_stock': total_network,
            'network_sizes': total_sizes,
            'wms_total': total_wms,
            'wms_sizes': wms_sizes_detail,
            'can_order_count': can_order_count,
            'is_broken_grid': is_broken_grid,
            'needs_schlopka': needs_schlopka,
            'days_no_sale': days,
        }
        if role != 'user':
            item['branch'] = r.get('branch', '')

        key = (art, r.get('branch', ''))
        if key not in seen:
            seen.add(key)
            result.append(item)

    return jsonify(result)


@app.route('/api/zero-sales/auto-schlopka', methods=['POST'])
@admin_required
def auto_schlopka():
    """
    Авто-схлопка: собирает артикулы с сетью < 50 шт и
    равномерно распределяет по 4 флагманам (по 1 шт каждого размера).
    Правильная сетка = по 1 штуке КАЖДОГО размера (не 3 одинаковых).
    """
    data = request.get_json() or {}
    articles = data.get('articles', [])

    if not articles:
        return jsonify({'error': 'Нет артикулов'}), 400

    FLAGMAN_LIST = ['TASHKENT CITY MALL', 'ALAYSKIY', 'ATLAS CHIMGAN', 'Shota Rustavely']

    conn = get_db(); cur = conn.cursor()

    # Все остатки по этим артикулам во всех филиалах
    # Фильтруем артикулы на скидке 70% — они только для аутлетов
    cur.execute("SELECT DISTINCT article FROM catalog WHERE discount > 0")
    discount_arts = {r['article'] for r in cur.fetchall()}
    articles = [a for a in articles if a not in discount_arts]

    if not articles:
        cur.close(); conn.close()
        return jsonify({'error': 'Все артикулы на скидке 70% — они только для аутлетов'}), 400

    cur.execute("""
        SELECT bs.article, bs.branch, bs.size, bs.qty,
               MIN(c.name) as name, MIN(c.category) as category
        FROM branch_stock bs
        JOIN catalog c ON c.article = bs.article
        WHERE bs.article = ANY(%s) AND bs.qty > 0
        AND c.discount = 0
        GROUP BY bs.article, bs.branch, bs.size, bs.qty
        ORDER BY bs.article, bs.size
    """, (articles,))
    rows = cur.fetchall()

    if not rows:
        cur.close(); conn.close()
        return jsonify({'error': 'Нет остатков'}), 400

    # Продажи флагманов за 30 дней по этим артикулам
    art_list_no_a = [a.rstrip('A') for a in articles]
    cur.execute("""
        SELECT article, branch, SUM(qty) as sold_30d
        FROM sales
        WHERE (article = ANY(%s) OR article = ANY(%s))
        AND branch = ANY(%s)
        AND sale_date >= CURRENT_DATE - 30
        GROUP BY article, branch
    """, (articles, art_list_no_a, FLAGMAN_LIST))
    sales_by_flagman = {}  # (art, branch) -> sold_30d
    for r in cur.fetchall():
        art_key = r['article'].rstrip('A') + 'A' if not r['article'].endswith('A') else r['article']
        sales_by_flagman[(art_key, r['branch'])] = int(r['sold_30d'] or 0)
        sales_by_flagman[(r['article'], r['branch'])] = int(r['sold_30d'] or 0)

    from collections import defaultdict
    art_sizes = defaultdict(lambda: defaultdict(dict))
    art_names = {}
    for r in rows:
        art_sizes[r['article']][r['size']][r['branch']] = int(r['qty'] or 0)
        art_names[r['article']] = r['name'] or r['category'] or ''

    # Исключаем из доноров Самарканд и Андижан
    NO_TAKE_BRANCHES = {'UZBEGIM ANDIJAN', 'Samarkand', 'SAMARKAND', 'М.БАРАКА САМАРКАНД'}

    schlopka_items = []  # (article, name, size, from_branch, to_flagman, qty)

    for art, sizes_data in art_sizes.items():
        name = art_names.get(art, '')
        all_sizes = sorted(sizes_data.keys())

        # Продажи флагманов по этому артикулу за 30 дней
        flagman_sales = {}
        total_flagman_sales = 0
        for f in FLAGMAN_LIST:
            sold = sales_by_flagman.get((art, f), 0)
            flagman_sales[f] = sold
            total_flagman_sales += sold

        # Если никто ничего не продавал — делим поровну
        if total_flagman_sales == 0:
            flagman_weights = {f: 1/len(FLAGMAN_LIST) for f in FLAGMAN_LIST}
        else:
            flagman_weights = {f: flagman_sales[f]/total_flagman_sales for f in FLAGMAN_LIST}

        # Считаем суммарный остаток который можно собрать (из всех доноров кроме NO_TAKE)
        # и сколько нужно дать каждому флагману
        for size in all_sizes:
            branch_stocks = sizes_data[size]

            # Доноры — не флагманы и не запрещённые филиалы
            donors = {b: q for b, q in branch_stocks.items()
                      if b not in FLAGMAN_LIST and b not in NO_TAKE_BRANCHES and q > 0}
            if not donors:
                continue

            total_available = sum(donors.values())
            if total_available == 0:
                continue

            # Сколько уже есть у флагманов
            flagman_has = {f: branch_stocks.get(f, 0) for f in FLAGMAN_LIST}

            # Распределяем пропорционально продажам
            # Сначала считаем целевое кол-во для каждого флагмана
            targets = {}
            for f in FLAGMAN_LIST:
                weight = flagman_weights[f]
                # Базовая норма: минимум 1 сетка, максимум пропорционально
                base = max(1, round(total_available * weight))
                targets[f] = base

            # Сортируем флагманов по продажам (самые продающие — первыми)
            sorted_flagmans = sorted(FLAGMAN_LIST, key=lambda f: flagman_sales[f], reverse=True)

            remaining_donors = dict(donors)
            for flagman in sorted_flagmans:
                current = flagman_has[flagman]
                target = targets[flagman]
                need = max(0, target - current)

                for _ in range(need):
                    if not remaining_donors:
                        break
                    best_donor = max(remaining_donors, key=lambda b: remaining_donors[b])
                    schlopka_items.append((art, name, size, best_donor, flagman, 1))
                    remaining_donors[best_donor] -= 1
                    if remaining_donors[best_donor] <= 0:
                        del remaining_donors[best_donor]

            # Остаток после флагманов — отдаём самому продающему
            if remaining_donors and sorted_flagmans:
                top_flagman = sorted_flagmans[0]
                for donor, qty in remaining_donors.items():
                    if qty > 0:
                        schlopka_items.append((art, name, size, donor, top_flagman, qty))

    if not schlopka_items:
        cur.close(); conn.close()
        return jsonify({'error': 'Флагманы уже имеют полные сетки по этим артикулам'}), 400

    # Создаём сессию схлопки
    session_name = f'Авто-схлопка {datetime.now().strftime("%d.%m.%Y %H:%M")}'
    fname = f'auto_schlopka_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    cur.execute(
        'INSERT INTO schlopka_sessions (name,filename,created_by) VALUES (%s,%s,%s) RETURNING id',
        (session_name, fname, session.get('username'))
    )
    sid = cur.fetchone()['id']

    # Группируем по from_branch (лист схлопки = откуда)
    from collections import Counter
    items_inserted = 0
    for art, name, size, from_branch, to_flagman, qty in schlopka_items:
        cur.execute(
            'INSERT INTO schlopka_items (session_id,article,name,size,branch,qty,note) VALUES (%s,%s,%s,%s,%s,%s,%s)',
            (sid, art, name, size, from_branch, qty, to_flagman)
        )
        items_inserted += 1

    conn.commit()

    # Summary per flagman
    flagman_totals = Counter(item[4] for item in schlopka_items)

    cur.close(); conn.close()
    return jsonify({
        'ok': True,
        'session_id': sid,
        'session_name': session_name,
        'items': items_inserted,
        'by_flagman': dict(flagman_totals),
        'articles': len(set(i[0] for i in schlopka_items))
    })


# ===== WAREHOUSE MORNING REPORT =====

@app.route('/api/warehouse/report')
@login_required
def warehouse_report():
    """
    Утренний отчёт склада:
    - Артикулы с остатком WMS < 10 шт (по размеру)
    - Артикулы с суммарным остатком по сети < 10 шт
    Исключает товары на скидке 70%
    """
    if session.get('role') not in ('admin', 'warehouse'):
        return jsonify({'error': 'Нет доступа'}), 403

    conn = get_db(); cur = conn.cursor()

    # 1. Мало на складе WMS (по каждому размеру < 10)
    cur.execute("""
        SELECT
            c.article,
            MIN(c.name) as name,
            MIN(c.abc) as abc,
            MIN(c.season) as season,
            MIN(c.category) as category,
            c.size,
            c.wms_stock,
            SUM(bs.qty) as network_stock
        FROM catalog c
        LEFT JOIN branch_stock bs ON bs.article = c.article AND bs.size = c.size
        WHERE c.wms_stock > 0 AND c.wms_stock < 10
        AND c.discount = 0
        GROUP BY c.article, c.name, c.abc, c.season, c.category, c.size, c.wms_stock
        ORDER BY c.wms_stock ASC, MIN(c.abc), c.article
        LIMIT 300
    """)
    low_wms_rows = cur.fetchall()

    # 2. Мало по сети (суммарный остаток < 10)
    cur.execute("""
        SELECT
            c.article,
            MIN(c.name) as name,
            MIN(c.abc) as abc,
            MIN(c.season) as season,
            MIN(c.category) as category,
            SUM(c.wms_stock) as total_wms,
            SUM(bs.qty) as total_network,
            COUNT(DISTINCT c.size) as size_count,
            array_agg(c.size ORDER BY c.size) as sizes,
            array_agg(c.wms_stock ORDER BY c.size) as wms_stocks
        FROM catalog c
        LEFT JOIN branch_stock bs ON bs.article = c.article
        WHERE c.discount = 0
        GROUP BY c.article
        HAVING SUM(c.wms_stock) + COALESCE(SUM(bs.qty), 0) < 10
        AND SUM(c.wms_stock) + COALESCE(SUM(bs.qty), 0) > 0
        ORDER BY MIN(c.abc), SUM(c.wms_stock) + COALESCE(SUM(bs.qty), 0) ASC
        LIMIT 200
    """)
    low_network_rows = cur.fetchall()

    # 3. Единички на складе (wms_stock = 1)
    cur.execute("""
        SELECT
            c.article,
            MIN(c.name) as name,
            MIN(c.abc) as abc,
            MIN(c.season) as season,
            MIN(c.category) as category,
            array_agg(c.size ORDER BY c.size) as sizes,
            array_agg(c.wms_stock ORDER BY c.size) as wms_stocks
        FROM catalog c
        WHERE c.discount = 0
        GROUP BY c.article
        HAVING SUM(CASE WHEN c.wms_stock = 1 THEN 1 ELSE 0 END) > 0
        AND MIN(c.abc) IN ('A', 'B')
        ORDER BY MIN(c.abc), c.article
        LIMIT 200
    """)
    singles_rows = cur.fetchall()

    cur.close(); conn.close()

    # Format results
    low_wms = []
    for r in low_wms_rows:
        low_wms.append({
            'article': r['article'],
            'name': r['name'] or '',
            'abc': r['abc'] or 'C',
            'season': r['season'] or '',
            'category': r['category'] or '',
            'size': r['size'],
            'wms_stock': int(r['wms_stock'] or 0),
            'network_stock': int(r['network_stock'] or 0),
        })

    low_network = []
    for r in low_network_rows:
        sizes_detail = []
        sizes = r['sizes'] or []
        stocks = r['wms_stocks'] or []
        for i, sz in enumerate(sizes):
            w = int(stocks[i] or 0) if i < len(stocks) else 0
            sizes_detail.append({'size': sz, 'wms': w})
        low_network.append({
            'article': r['article'],
            'name': r['name'] or '',
            'abc': r['abc'] or 'C',
            'season': r['season'] or '',
            'category': r['category'] or '',
            'total_wms': int(r['total_wms'] or 0),
            'total_network': int(r['total_network'] or 0),
            'sizes': sizes_detail,
        })

    singles = []
    for r in singles_rows:
        sizes = r['sizes'] or []
        stocks = r['wms_stocks'] or []
        single_sizes = [{'size': sizes[i], 'wms': int(stocks[i] or 0)}
                        for i in range(len(sizes)) if i < len(stocks) and int(stocks[i] or 0) == 1]
        if single_sizes:
            singles.append({
                'article': r['article'],
                'name': r['name'] or '',
                'abc': r['abc'] or 'C',
                'season': r['season'] or '',
                'category': r['category'] or '',
                'single_sizes': single_sizes,
            })

    return jsonify({
        'low_wms': low_wms,
        'low_network': low_network,
        'singles': singles,
        'generated_at': datetime.now().strftime('%d.%m.%Y %H:%M'),
        'summary': {
            'low_wms_count': len(low_wms),
            'low_network_count': len(low_network),
            'singles_count': len(singles),
        }
    })


# ===== POWER BI =====

def _pbi_export_file():
    """Internal: export Power BI report and return raw bytes"""
    import urllib.request, json as _json, time
    token = get_pbi_token()
    if not token:
        raise Exception('PBI_CLIENT_SECRET не настроен в Railway Variables')

    # Start export job
    url = f'https://api.powerbi.com/v1.0/myorg/groups/me/reports/{PBI_REPORT_ID}/ExportTo'
    payload = _json.dumps({'format': 'XLSX'}).encode()
    req = urllib.request.Request(url, data=payload, method='POST',
        headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'})
    with urllib.request.urlopen(req, timeout=30) as r:
        job = _json.loads(r.read())

    export_id = job.get('id')
    if not export_id:
        raise Exception(f'Не удалось запустить экспорт: {job}')

    # Poll for completion (max 90 sec)
    for _ in range(18):
        time.sleep(5)
        poll_url = f'https://api.powerbi.com/v1.0/myorg/groups/me/reports/{PBI_REPORT_ID}/exports/{export_id}'
        req2 = urllib.request.Request(poll_url, headers={'Authorization': f'Bearer {token}'})
        with urllib.request.urlopen(req2, timeout=15) as r2:
            status = _json.loads(r2.read())
        if status.get('status') == 'Succeeded':
            break
        if status.get('status') == 'Failed':
            raise Exception(f'Экспорт не удался: {status}')

    # Download file
    dl_url = f'https://api.powerbi.com/v1.0/myorg/groups/me/reports/{PBI_REPORT_ID}/exports/{export_id}/file'
    req3 = urllib.request.Request(dl_url, headers={'Authorization': f'Bearer {token}'})
    with urllib.request.urlopen(req3, timeout=60) as r3:
        return r3.read()


@app.route('/api/powerbi/export')
@admin_required
def powerbi_export():
    """Export Power BI report to Excel for download"""
    try:
        file_data = _pbi_export_file()
        fname = f'PowerBI_Report_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return send_file(
            io.BytesIO(file_data),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=fname
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/powerbi/sync', methods=['POST'])
@admin_required
def powerbi_sync():
    """Export from Power BI and immediately import as sales data"""
    import re as _re
    replace = request.json.get('replace', False) if request.is_json else False
    try:
        file_data = _pbi_export_file()
        wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # Find header row
        header_idx = None
        for i, row in enumerate(rows[:10]):
            row_s = [str(c).strip() if c else '' for c in row]
            if any('артикул' in v.lower() or 'article' in v.lower() for v in row_s):
                header_idx = i; break
        if header_idx is None:
            return jsonify({'error': 'Не найден заголовок в файле Power BI'}), 400

        h = [str(c).strip() if c else '' for c in rows[header_idx]]
        season_col = art_col = cat_col = branch_col = ref_col = qty_col = price_col = amount_col = None
        for j, v in enumerate(h):
            vl = v.lower()
            if 'сезон' in vl or 'season' in vl: season_col = j
            if 'артикул' in vl or 'article' in vl: art_col = j
            if 'вид' in vl or 'category' in vl or 'type' in vl: cat_col = j
            if 'магазин' in vl or 'branch' in vl or 'store' in vl: branch_col = j
            if 'ссылка' in vl or 'ref' in vl or 'чек' in vl: ref_col = j
            if 'количество' in vl or 'qty' in vl or 'кол' in vl: qty_col = j
            if 'цена' in vl or 'price' in vl: price_col = j
            if 'сумма' in vl or 'amount' in vl or 'выручка' in vl: amount_col = j

        conn = get_db(); cur = conn.cursor()
        if replace:
            cur.execute('DELETE FROM sales')

        inserted = 0
        for row in rows[header_idx+1:]:
            if not row: continue
            art = str(row[art_col]).strip() if art_col is not None and row[art_col] else ''
            if not art or art in ('None','nan',''): continue
            season = str(row[season_col]).strip() if season_col is not None and row[season_col] else ''
            cat = str(row[cat_col]).strip() if cat_col is not None and row[cat_col] else ''
            branch = str(row[branch_col]).strip() if branch_col is not None and row[branch_col] else ''
            ref = str(row[ref_col]).strip() if ref_col is not None and row[ref_col] else ''
            try: qty = int(float(str(row[qty_col]).replace(' ','').replace(' ',''))) if qty_col is not None and row[qty_col] and str(row[qty_col]) not in ('None','nan') else 1
            except: qty = 1
            try: price = float(str(row[price_col]).replace(' ','').replace(' ','').replace(',','.')) if price_col is not None and row[price_col] and str(row[price_col]) not in ('None','nan') else 0
            except: price = 0
            try: amount = float(str(row[amount_col]).replace(' ','').replace(' ','').replace(',','.')) if amount_col is not None and row[amount_col] and str(row[amount_col]) not in ('None','nan') else 0
            except: amount = 0
            sale_date = None
            m = _re.search(r'от (\d{2})\.(\d{2})\.(\d{4})', ref)
            if m:
                try: sale_date = f'{m.group(3)}-{m.group(2)}-{m.group(1)}'
                except: pass
            cur.execute(
                'INSERT INTO sales (season,article,category,branch,sale_date,qty,price,amount) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)',
                (season, art.rstrip('A'), cat, branch, sale_date, qty, price, amount)
            )
            inserted += 1

        conn.commit(); cur.close(); conn.close()
        return jsonify({'ok': True, 'inserted': inserted, 'replaced': replace})
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'tb': traceback.format_exc()}), 500


@app.route('/api/warehouse/report/excel')
@login_required
def warehouse_report_excel():
    """Export warehouse report to Excel"""
    if session.get('role') not in ('admin', 'warehouse'):
        return jsonify({'error': 'Нет доступа'}), 403

    conn = get_db(); cur = conn.cursor()

    # Singles A/B
    cur.execute("""
        SELECT c.article, MIN(c.name) as name, MIN(c.abc) as abc,
               MIN(c.season) as season, MIN(c.category) as category,
               array_agg(c.size ORDER BY c.size) as sizes,
               array_agg(c.wms_stock ORDER BY c.size) as wms_stocks
        FROM catalog c WHERE c.discount = 0
        GROUP BY c.article
        HAVING SUM(CASE WHEN c.wms_stock = 1 THEN 1 ELSE 0 END) > 0
        AND MIN(c.abc) IN ('A', 'B')
        ORDER BY MIN(c.abc), c.article LIMIT 300
    """)
    singles = cur.fetchall()

    # Low WMS < 10
    cur.execute("""
        SELECT c.article, MIN(c.name) as name, MIN(c.abc) as abc,
               MIN(c.season) as season, MIN(c.category) as category,
               c.size, c.wms_stock
        FROM catalog c
        WHERE c.wms_stock > 0 AND c.wms_stock < 10 AND c.discount = 0
        GROUP BY c.article, c.name, c.abc, c.season, c.category, c.size, c.wms_stock
        ORDER BY c.wms_stock ASC, MIN(c.abc), c.article LIMIT 300
    """)
    low_wms = cur.fetchall()

    # Low network < 10
    cur.execute("""
        SELECT c.article, MIN(c.name) as name, MIN(c.abc) as abc,
               MIN(c.season) as season, MIN(c.category) as category,
               SUM(c.wms_stock) as total_wms,
               array_agg(c.size ORDER BY c.size) as sizes,
               array_agg(c.wms_stock ORDER BY c.size) as wms_stocks
        FROM catalog c LEFT JOIN branch_stock bs ON bs.article = c.article
        WHERE c.discount = 0
        GROUP BY c.article
        HAVING SUM(c.wms_stock) + COALESCE(SUM(bs.qty), 0) < 10
        AND SUM(c.wms_stock) + COALESCE(SUM(bs.qty), 0) > 0
        ORDER BY MIN(c.abc), SUM(c.wms_stock) ASC LIMIT 200
    """)
    low_network = cur.fetchall()
    cur.close(); conn.close()

    from openpyxl.utils import get_column_letter as gcl
    wb = openpyxl.Workbook()

    hdr_fill = PatternFill('solid', fgColor='1A1A2E')
    amber_fill = PatternFill('solid', fgColor='FFF3E0')
    red_fill = PatternFill('solid', fgColor='FFEBEE')
    green_fill = PatternFill('solid', fgColor='E8F5E9')
    thin = Side(style='thin', color='DDDDDD')
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal='center', vertical='center')
    lft = Alignment(horizontal='left', vertical='center')

    def hdr_cell(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = hdr_fill; c.font = Font(bold=True, color='FFFFFF', size=10)
        c.alignment = ctr; c.border = brd
        return c

    # Sheet 1: Singles
    ws1 = wb.active; ws1.title = '⚠️ Единички A-B'
    ws1.row_dimensions[1].height = 30
    for ci, h in enumerate(['Артикул','Название','ABC','Сезон','Категория','Размер','WMS'], 1):
        hdr_cell(ws1, 1, ci, h)

    row_idx = 2
    for r in singles:
        sizes = r['sizes'] or []
        stocks = r['wms_stocks'] or []
        single_sizes = [(sizes[i], int(stocks[i] or 0)) for i in range(len(sizes)) if i < len(stocks) and int(stocks[i] or 0) == 1]
        for sz, wms in single_sizes:
            vals = [r['article'], r['name'], r['abc'], r['season'], r['category'], sz, wms]
            for ci, v in enumerate(vals, 1):
                cell = ws1.cell(row=row_idx, column=ci, value=v)
                cell.border = brd; cell.alignment = ctr if ci >= 3 else lft
                cell.font = Font(size=9)
                if ci == 7: cell.fill = amber_fill; cell.font = Font(size=9, bold=True, color='E65100')
            row_idx += 1

    for ci, w in enumerate([14,36,6,10,12,8,8], 1):
        ws1.column_dimensions[gcl(ci)].width = w
    ws1.freeze_panes = 'A2'

    # Sheet 2: Low WMS
    ws2 = wb.create_sheet('📦 Мало на WMS')
    ws2.row_dimensions[1].height = 30
    for ci, h in enumerate(['Артикул','Название','ABC','Сезон','Категория','Размер','WMS остаток'], 1):
        hdr_cell(ws2, 1, ci, h)

    for ri, r in enumerate(low_wms, 2):
        vals = [r['article'], r['name'], r['abc'], r['season'], r['category'], r['size'], int(r['wms_stock'] or 0)]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.border = brd; cell.alignment = ctr if ci >= 3 else lft; cell.font = Font(size=9)
            if ci == 7:
                wms = int(r['wms_stock'] or 0)
                cell.fill = red_fill if wms <= 3 else amber_fill
                cell.font = Font(size=9, bold=True, color='B71C1C' if wms <= 3 else 'E65100')

    for ci, w in enumerate([14,36,6,10,12,8,10], 1):
        ws2.column_dimensions[gcl(ci)].width = w
    ws2.freeze_panes = 'A2'

    # Sheet 3: Low Network
    ws3 = wb.create_sheet('🌐 Мало в сети')
    ws3.row_dimensions[1].height = 30
    for ci, h in enumerate(['Артикул','Название','ABC','Сезон','Категория','WMS всего','Размеры (WMS)'], 1):
        hdr_cell(ws3, 1, ci, h)

    for ri, r in enumerate(low_network, 2):
        sizes = r['sizes'] or []
        stocks = r['wms_stocks'] or []
        sizes_str = ', '.join(f"{sizes[i]}:{int(stocks[i] or 0)}" for i in range(len(sizes)) if i < len(stocks))
        vals = [r['article'], r['name'], r['abc'], r['season'], r['category'], int(r['total_wms'] or 0), sizes_str]
        for ci, v in enumerate(vals, 1):
            cell = ws3.cell(row=ri, column=ci, value=v)
            cell.border = brd; cell.alignment = ctr if ci >= 3 else lft; cell.font = Font(size=9)
            if ci == 6:
                cell.fill = red_fill if int(r['total_wms'] or 0) <= 3 else amber_fill
                cell.font = Font(size=9, bold=True)

    for ci, w in enumerate([14,36,6,10,12,10,40], 1):
        ws3.column_dimensions[gcl(ci)].width = w
    ws3.freeze_panes = 'A2'

    output = io.BytesIO(); wb.save(output); output.seek(0)
    fname = f'Warehouse_Report_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
